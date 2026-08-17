#!/usr/bin/env python3
"""
Add statistics sheets to the Excel report based on instructions:
1. Q38-Q56 Statistics by PDN code - trait ratings distribution (1-7 scale)
2. Q57-Q61 Statistics by PDN code - first choice preferences
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import json

# Load questions definitions
with open('/Users/tomer.gur/dev-tools/pdn_chat/app/data/questions.json', 'r', encoding='utf-8') as f:
    questions_data = json.load(f)

# Question text mapping for Q38-Q56
Q38_Q56_QUESTIONS = {
    38: "באיזו מידה אתה נוטה לבדוק לעומק לפני נטילת סיכון?",
    39: "עד כמה קשה לך להסתדר עם מצבי חוסר ודאות?",
    40: "באיזו מידה אתה מקבל החלטות במהירות באופן כללי?",
    41: "האם קשה לך לקבל הנחיות והובלה מאחרים?",
    42: "עד כמה אתה מרגיש נוח להוביל אחרים כשאין הנהגה ברורה?",
    43: "מוחצנות/מופנמות בילדות",
    44: "מוחצנות/מופנמות בבגרות",
    45: "הובלה/הצטרפות בילדות",
    46: "הובלה/הצטרפות בבגרות",
    47: "דברנות/שתקנות בילדות",
    48: "דברנות/שתקנות בבגרות",
    49: "עימותים/ריצוי בילדות",
    50: "עימותים/ריצוי בבגרות",
    51: "עמידה בזמנים בילדות",
    52: "עמידה בזמנים בבגרות",
    53: "סדר/בלגן בילדות",
    54: "סדר/בלגן בבגרות",
    55: "מאופקות/נועזות בילדות",
    56: "מאופקות/נועזות בבגרות",
}

# Trait codes for Q38-Q42
Q38_42_TRAITS = {
    38: ("A", "T"),  # A=לא מוודא, T=מוודא
    39: ("A", "T"),  # A=לא מתקשה, T=מתקשה
    40: ("P", "E"),  # P=מתקשה להחליט, E=מחליט מהר
    41: ("P", "E"),  # P=אין קושי, E=קשה
    42: ("P", "E"),  # P=נשאר בצד, E=לוקח הובלה
}

# Trait codes for Q43-Q56
Q43_56_TRAITS = {
    43: ("TP", "AE"), 44: ("TP", "AE"), 45: ("TP", "AE"), 46: ("TP", "AE"),
    47: ("TP", "AE"), 48: ("TP", "AE"), 49: ("TP", "AE"), 50: ("TP", "AE"),
    51: ("AP", "TE"), 52: ("AP", "TE"), 53: ("AP", "TE"), 54: ("AP", "TE"),
    55: ("TP", "AE"), 56: ("TP", "AE"),
}

# PDN codes order
PDN_CODES = ["E1", "E5", "E9", "A3", "A7", "A11", "T4", "T8", "T12", "P2", "P6", "P10"]

# Styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def extract_pdn_code(user_str):
    """Extract PDN code from user string like 'P10 | UID2EE4D3 | Tomer Gur'"""
    if not user_str:
        return None
    parts = user_str.split('|')
    if parts:
        code = parts[0].strip()
        if code in PDN_CODES:
            return code
    return None

def parse_rating(value):
    """Parse rating value - could be number 1-7 or code like T, A, P, E, TP, AE, etc."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    # For non-numeric values, return as-is (it's a code selection)
    return str(value)

def load_data():
    """Load data from the Excel file"""
    wb = openpyxl.load_workbook('/Users/tomer.gur/dev-tools/pdn_chat/docs/combined_matrix_report.xlsx')
    ws = wb['תשובות משתמשים']
    
    # Get header
    header = [cell.value for cell in ws[1]]
    
    # Find column indices
    q38_idx = header.index('Q38')
    q56_idx = header.index('Q56')
    q57_idx = header.index('Q57')
    q61_idx = header.index('Q61')
    
    # Collect data by PDN code
    data_by_code = defaultdict(lambda: {
        'q38_q56': defaultdict(list),  # question -> list of answers
        'q57_q61': defaultdict(list),  # question -> list of first choices
        'count': 0
    })
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        user = row[0]
        pdn_code = extract_pdn_code(user)
        if not pdn_code:
            continue
        
        data_by_code[pdn_code]['count'] += 1
        
        # Q38-Q56 (ratings 1-7)
        for i, q_num in enumerate(range(38, 57)):
            val = row[q38_idx + i]
            if val is not None:
                data_by_code[pdn_code]['q38_q56'][q_num].append(parse_rating(val))
        
        # Q57-Q61 (first choice from ranking)
        for i, q_num in enumerate(range(57, 62)):
            val = row[q57_idx + i]
            if val is not None and val not in ['קופון הטבה/חוויה', 'מקצועיות', None]:
                data_by_code[pdn_code]['q57_q61'][q_num].append(str(val))
    
    return wb, data_by_code

def create_q38_q56_sheet(wb, data_by_code):
    """Create statistics sheet for Q38-Q56 by PDN code"""
    if 'Q38-Q56 סטטיסטיקה' in wb.sheetnames:
        del wb['Q38-Q56 סטטיסטיקה']
    
    ws = wb.create_sheet('Q38-Q56 סטטיסטיקה')
    
    # Header row
    headers = ['שאלה'] + [f'{code}\n({data_by_code[code]["count"]})' for code in PDN_CODES]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Data rows
    row_num = 2
    for q_num in range(38, 57):
        q_text = Q38_Q56_QUESTIONS.get(q_num, f'Q{q_num}')
        ws.cell(row=row_num, column=1, value=f'Q{q_num}\n{q_text}').border = thin_border
        ws.cell(row=row_num, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        
        for col, code in enumerate(PDN_CODES, 2):
            answers = data_by_code[code]['q38_q56'].get(q_num, [])
            if not answers:
                ws.cell(row=row_num, column=col, value='-').border = thin_border
            else:
                # Count occurrences of each answer
                counts = defaultdict(int)
                for ans in answers:
                    counts[ans] += 1
                
                # Format as percentage breakdown
                total = len(answers)
                parts = []
                for ans, cnt in sorted(counts.items(), key=lambda x: -x[1]):
                    pct = cnt * 100 // total
                    parts.append(f'{ans}: {pct}%')
                
                cell = ws.cell(row=row_num, column=col, value='\n'.join(parts))
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
        
        row_num += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    for col in range(2, len(PDN_CODES) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    return ws

def create_q57_q61_sheet(wb, data_by_code):
    """Create statistics sheet for Q57-Q61 by PDN code (first choice percentages)"""
    if 'Q57-Q61 סטטיסטיקה' in wb.sheetnames:
        del wb['Q57-Q61 סטטיסטיקה']
    
    ws = wb.create_sheet('Q57-Q61 סטטיסטיקה')
    
    # Header row
    headers = ['שאלה'] + [f'{code}\n({data_by_code[code]["count"]})' for code in PDN_CODES]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Question descriptions
    q57_61_desc = {
        57: "דירוג תכונות (אסרטיבי/אכפתי/שיטתי/שופע רעיונות)",
        58: "דירוג תכונות (מוביל/תומך/זהיר/רעיוניסט)",
        59: "דירוג תכונות (ממוקד מטרה/נותן/שיטתי/אופטימי)",
        60: "מה היית רוצה לקבל מהאבחון",
        61: "מה חשוב לך שהאבחון יספק"
    }
    
    # Data rows
    row_num = 2
    for q_num in range(57, 62):
        q_text = q57_61_desc.get(q_num, f'Q{q_num}')
        ws.cell(row=row_num, column=1, value=f'Q{q_num}\n{q_text}').border = thin_border
        ws.cell(row=row_num, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        
        for col, code in enumerate(PDN_CODES, 2):
            first_choices = data_by_code[code]['q57_q61'].get(q_num, [])
            if not first_choices:
                ws.cell(row=row_num, column=col, value='-').border = thin_border
            else:
                # Count first choice occurrences
                counts = defaultdict(int)
                for choice in first_choices:
                    counts[choice] += 1
                
                # Format as percentage breakdown for T/A/E/P
                total = len(first_choices)
                parts = []
                for trait in ['T', 'A', 'E', 'P']:
                    if trait in counts:
                        pct = counts[trait] * 100 // total
                        parts.append(f'{trait}: {pct}%')
                
                # Add any other values
                for choice, cnt in counts.items():
                    if choice not in ['T', 'A', 'E', 'P']:
                        pct = cnt * 100 // total
                        parts.append(f'{choice[:10]}: {pct}%')
                
                cell = ws.cell(row=row_num, column=col, value='\n'.join(parts) if parts else '-')
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
        
        row_num += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    for col in range(2, len(PDN_CODES) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    return ws

def create_organized_answers_sheet(wb):
    """Create organized answers sheet with Name | UID | Verified PDN Code"""
    if 'גיליון תשובות מאורגנים' in wb.sheetnames:
        del wb['גיליון תשובות מאורגנים']
    
    ws_source = wb['תשובות משתמשים']
    ws = wb.create_sheet('גיליון תשובות מאורגנים', 0)  # Insert at beginning
    
    # Headers
    headers = ['Name', 'UID', 'Verified PDN Code']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Parse data from source
    row_num = 2
    for row in ws_source.iter_rows(min_row=2, values_only=True):
        user_str = row[0]
        if not user_str:
            continue
        
        # Parse "P10 | UID2EE4D3 | Tomer Gur"
        parts = [p.strip() for p in user_str.split('|')]
        if len(parts) >= 3:
            pdn_code = parts[0]
            uid = parts[1].replace('UID', '')
            name = parts[2]
            
            ws.cell(row=row_num, column=1, value=name).border = thin_border
            ws.cell(row=row_num, column=2, value=uid).border = thin_border
            ws.cell(row=row_num, column=3, value=pdn_code).border = thin_border
            row_num += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    
    return ws

def main():
    print("Loading data from Excel...")
    wb, data_by_code = load_data()
    
    print("\nData summary:")
    for code in PDN_CODES:
        count = data_by_code[code]['count']
        if count > 0:
            print(f"  {code}: {count} users")
    
    print("\nCreating organized answers sheet...")
    create_organized_answers_sheet(wb)
    
    print("Creating Q38-Q56 statistics sheet...")
    create_q38_q56_sheet(wb, data_by_code)
    
    print("Creating Q57-Q61 statistics sheet...")
    create_q57_q61_sheet(wb, data_by_code)
    
    # Save
    output_path = '/Users/tomer.gur/dev-tools/pdn_chat/docs/combined_matrix_report.xlsx'
    wb.save(output_path)
    print(f"\nSaved to: {output_path}")
    print(f"Sheets: {wb.sheetnames}")

if __name__ == '__main__':
    main()
