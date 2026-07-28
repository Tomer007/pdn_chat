#!/usr/bin/env python3
"""
Combined PDN Matrix Report

Creates a single HTML report with 3 sections:
1. User Answers Matrix - rows: users (UID), columns: questions, values: answers
2. PDN Code Statistics Matrix - rows: questions, columns: 12 PDN codes, values: aggregate stats
3. Insights - question significance (מובהקות) analysis
"""

import csv
import json
import os
import sys
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import math

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

SAVED_RESULTS_DIR = PROJECT_ROOT / "saved_results"
QUESTIONS_FILE = PROJECT_ROOT / "app" / "data" / "questions.json"
USER_METADATA_CSV = SAVED_RESULTS_DIR / "user_metadata.csv"
TEST_USERS_FILE = PROJECT_ROOT / "app" / "data" / "test_users.json"

PDN_12_CODES = ["E1", "E5", "E9", "A3", "A7", "A11", "T4", "T8", "T12", "P2", "P6", "P10"]


# --- Data Loading ---

def load_test_emails():
    """Load test user emails from test_users.json."""
    if not TEST_USERS_FILE.exists():
        return set()
    try:
        with open(TEST_USERS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return set(e.lower() for e in data.get("test_emails", []))
    except (json.JSONDecodeError, IOError):
        return set()

def load_questions():
    with open(QUESTIONS_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    questions = {}
    for phase_name, phase_data in data.get("phases", {}).items():
        if phase_name == "PersonalDetails":
            continue
        if phase_name == "PartF":
            continue  # Skip PartF - marketing/preference questions, not PDN relevant
        for q_num, q_data in phase_data.get("questions", {}).items():
            codes = set()
            for opt in q_data.get("options", []):
                if opt.get("code"):
                    codes.add(opt["code"])
            codes_str = "/".join(sorted(codes)) if codes else "-"
            questions[q_num] = {
                "text": q_data.get("text", ""),
                "codes": codes_str,
                "phase": phase_name,
                "options": q_data.get("options", []),
            }
    return questions


def load_users(exclude_test=True):
    test_emails = load_test_emails() if exclude_test else set()
    users = []
    excluded_count = 0
    no_diag_count = 0
    with open(USER_METADATA_CSV, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            email = row.get("Email", "").strip()
            if email.lower() in test_emails:
                excluded_count += 1
                continue
            # Use Diagnose PDN Code (validated by diagnoser) instead of system PDN Code
            pdn_code = row.get("Diagnose PDN Code", "").strip()
            # Only include users with valid diagnosed PDN codes (one of the 12)
            if pdn_code not in PDN_12_CODES:
                no_diag_count += 1
                continue
            users.append({
                "uid": row.get("User ID", ""),
                "email": email,
                "pdn_code": pdn_code,
                "date": row.get("Date", ""),
            })
    if excluded_count:
        print(f"  Filtered out {excluded_count} test users (from app/data/test_users.json)")
    if no_diag_count:
        print(f"  Filtered out {no_diag_count} users without valid diagnoser code")
    return users


def get_user_dir(email):
    safe_username = "".join(c for c in email if c.isalnum() or c in (' ', '-', '_')).rstrip()
    safe_username = safe_username.replace(' ', '_')
    return SAVED_RESULTS_DIR / safe_username


def load_user_answers(email):
    user_dir = get_user_dir(email)
    answers_file = user_dir / f"{email}_answers.json"
    if not answers_file.exists():
        return None
    try:
        with open(answers_file, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return None


def _is_rank_order(q_num):
    """
    Returns True for questions where ranking values represent ORDER (1=best, higher=worse).
    PartB (Q27-37) and PartE (Q57-61) use drag-to-rank: rank 1 = most preferred.
    Returns False for PartC/PartD (Q38-56) which use a 0-12 scale: higher = stronger match.
    """
    try:
        n = int(q_num)
        return (27 <= n <= 37) or (57 <= n <= 61)
    except (ValueError, TypeError):
        return False


def format_answer_code(answer_data, q_num=None):
    if answer_data is None:
        return None
    if isinstance(answer_data, dict):
        if "selected_option_code" in answer_data:
            return answer_data["selected_option_code"]
        if "ranking" in answer_data:
            ranking = answer_data["ranking"]
            if isinstance(ranking, dict):
                if _is_rank_order(q_num):
                    # rank=1 is top choice -> pick LOWEST value
                    sorted_items = sorted(ranking.items(), key=lambda x: x[1])
                else:
                    # scale 0-12: higher = stronger identification -> pick HIGHEST value
                    sorted_items = sorted(ranking.items(), key=lambda x: -x[1])
                if sorted_items:
                    return sorted_items[0][0]
    return None


def get_answer_text(answer_data, q_num=None):
    if answer_data is None:
        return None
    if isinstance(answer_data, dict):
        if "selected_option_code" in answer_data:
            code = answer_data["selected_option_code"]
            for opt in answer_data.get("question_options", []):
                if opt.get("code") == code:
                    return opt.get("text", "").strip()
            return code
        if "ranking" in answer_data:
            ranking = answer_data["ranking"]
            if isinstance(ranking, dict):
                if _is_rank_order(q_num):
                    sorted_items = sorted(ranking.items(), key=lambda x: x[1])
                else:
                    sorted_items = sorted(ranking.items(), key=lambda x: -x[1])
                if sorted_items:
                    top_code = sorted_items[0][0]
                    for opt in answer_data.get("question_options", []):
                        if opt.get("code") == top_code:
                            return opt.get("text", "").strip()
                    return top_code
    return None


def format_answer_display(answer_data, q_num=None):
    """Format answer for the user matrix (code + short text)."""
    if answer_data is None:
        return ""
    if isinstance(answer_data, dict):
        if "selected_option_code" in answer_data:
            return answer_data["selected_option_code"]
        if "ranking" in answer_data:
            ranking = answer_data["ranking"]
            if isinstance(ranking, dict):
                if _is_rank_order(q_num):
                    sorted_items = sorted(ranking.items(), key=lambda x: x[1])
                else:
                    sorted_items = sorted(ranking.items(), key=lambda x: -x[1])
                return " ".join(f"{k}:{v}" for k, v in sorted_items)
    return ""


# --- Statistics ---

def compute_statistics(users, questions):
    users_by_code = defaultdict(list)
    for user in users:
        if user["pdn_code"] in PDN_12_CODES:
            users_by_code[user["pdn_code"]].append(user)

    all_answers = {}
    user_names = {}  # email -> (first_name, last_name)
    invalid_count = 0
    for user in users:
        answers = load_user_answers(user["email"])
        if answers:
            # Extract name from metadata
            meta = answers.get("metadata", {})
            first = meta.get("first_name", "")
            last = meta.get("last_name", "")
            user_names[user["email"]] = (first, last)

            # Filter out invalid answers - keep only valid question keys with proper structure
            answers_only = {}
            for k, v in answers.items():
                if k == "metadata":
                    continue
                # Valid answer must be a dict with selected_option_code or ranking
                if not isinstance(v, dict):
                    continue
                if "selected_option_code" not in v and "ranking" not in v:
                    continue
                # Validate selected_option_code is not empty/None
                if "selected_option_code" in v:
                    code = v["selected_option_code"]
                    if code is None or (isinstance(code, str) and code.strip() == ""):
                        continue
                # Validate ranking is a non-empty dict
                if "ranking" in v and "selected_option_code" not in v:
                    ranking = v["ranking"]
                    if not isinstance(ranking, dict) or len(ranking) == 0:
                        continue
                answers_only[k] = v

            if answers_only:
                all_answers[user["email"]] = answers_only
            else:
                invalid_count += 1
        else:
            invalid_count += 1

    if invalid_count:
        print(f"  Filtered out {invalid_count} users with no valid answers")

    stats = {}
    answer_texts = {}
    for q_num in questions:
        stats[q_num] = {}
        if q_num not in answer_texts:
            answer_texts[q_num] = {}

        for pdn_code in PDN_12_CODES:
            answer_counts = defaultdict(int)
            total = 0
            for user in users_by_code.get(pdn_code, []):
                user_answers = all_answers.get(user["email"], {})
                answer = user_answers.get(str(q_num))
                if answer is not None:
                    val = format_answer_code(answer, q_num)
                    if val is not None:
                        answer_counts[val] += 1
                        total += 1
                        if val not in answer_texts[q_num]:
                            txt = get_answer_text(answer, q_num)
                            if txt:
                                answer_texts[q_num][val] = txt
            stats[q_num][pdn_code] = {
                "counts": dict(answer_counts),
                "total": total,
            }

    return stats, users_by_code, all_answers, answer_texts, user_names


def compute_significance(stats, questions, users_by_code):
    """
    Compute significance (מובהקות) for each question.
    Measures how much the answer distribution differs across PDN codes.
    Uses chi-squared-like approach: if all codes answer the same, significance is low.
    If different codes answer differently, significance is high.
    """
    insights = []

    for q_num, q_data in questions.items():
        q_stats = stats.get(q_num, {})

        # Collect all answer codes used across all PDN codes
        all_answer_codes = set()
        total_respondents = 0
        for pdn_code in PDN_12_CODES:
            s = q_stats.get(pdn_code, {"counts": {}, "total": 0})
            all_answer_codes.update(s["counts"].keys())
            total_respondents += s["total"]

        if total_respondents < 5 or len(all_answer_codes) < 2:
            insights.append({
                "q_num": q_num,
                "text": q_data["text"],
                "codes": q_data["codes"],
                "phase": q_data["phase"],
                "significance": 0,
                "total_respondents": total_respondents,
                "insight": "לא מספיק נתונים",
                "dominant_by_code": {},
            })
            continue

        # Calculate overall distribution
        overall_counts = defaultdict(int)
        for pdn_code in PDN_12_CODES:
            s = q_stats.get(pdn_code, {"counts": {}, "total": 0})
            for ans, cnt in s["counts"].items():
                overall_counts[ans] += cnt

        overall_total = sum(overall_counts.values())
        if overall_total == 0:
            continue

        # Chi-squared-like significance
        chi_sq = 0
        codes_with_data = 0
        dominant_by_code = {}

        for pdn_code in PDN_12_CODES:
            s = q_stats.get(pdn_code, {"counts": {}, "total": 0})
            if s["total"] < 2:
                continue
            codes_with_data += 1

            # Find dominant answer for this code
            if s["counts"]:
                dom_ans = max(s["counts"], key=s["counts"].get)
                dom_pct = round(s["counts"][dom_ans] / s["total"] * 100)
                dominant_by_code[pdn_code] = f"{dom_ans} ({dom_pct}%)"

            for ans in all_answer_codes:
                observed = s["counts"].get(ans, 0)
                expected_rate = overall_counts.get(ans, 0) / overall_total
                expected = expected_rate * s["total"]
                if expected > 0:
                    chi_sq += (observed - expected) ** 2 / expected

        # Normalize significance to 0-100 scale
        if codes_with_data > 1:
            df = (codes_with_data - 1) * (len(all_answer_codes) - 1)
            if df > 0:
                significance = min(100, round(chi_sq / df * 20))
            else:
                significance = 0
        else:
            significance = 0

        # Generate insight text
        if significance >= 70:
            insight = "מובהקות גבוהה - שאלה מבדלת מאוד בין הקודים"
        elif significance >= 40:
            insight = "מובהקות בינונית - יש הבדלים בין הקודים"
        elif significance >= 20:
            insight = "מובהקות נמוכה - הבדלים קלים"
        else:
            insight = "לא מבדל - כל הקודים עונים דומה"

        insights.append({
            "q_num": q_num,
            "text": q_data["text"],
            "codes": q_data["codes"],
            "phase": q_data["phase"],
            "significance": significance,
            "total_respondents": total_respondents,
            "insight": insight,
            "dominant_by_code": dominant_by_code,
        })

    # Sort by significance descending
    insights.sort(key=lambda x: -x["significance"])
    return insights


# --- HTML Generation ---

def generate_html(users, questions, all_answers, stats, users_by_code, answer_texts, insights, user_names):
    def sort_key(q_num):
        try:
            return int(q_num)
        except ValueError:
            return 999

    sorted_q_nums = sorted(questions.keys(), key=sort_key)
    users_with_answers = [u for u in users if u["email"] in all_answers and all_answers[u["email"]]]
    code_counts = {code: len(ul) for code, ul in users_by_code.items()}

    # Build user list for the filter
    user_options_html = '<option value="all" selected>הכל</option>\n'
    for user in users_with_answers:
        email = user["email"]
        uid = user["uid"]
        pdn = user["pdn_code"] or "NA"
        first_name, last_name = user_names.get(email, ("", ""))
        display_name = f"{first_name} {last_name}".strip()
        label = f"{pdn} | {uid} | {display_name}" if display_name else f"{pdn} | {uid}"
        user_options_html += f'<option value="{uid}">{label}</option>\n'

    html = f"""<!DOCTYPE html>
<html dir="rtl" lang="he">
<head>
<meta charset="UTF-8">
<title>PDN - דוח מטריצות משולב (קוד מאבחן)</title>
<style>
* {{ box-sizing: border-box; }}
body {{
    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    margin: 0;
    padding: 20px;
    background: #f0f2f5;
    direction: rtl;
}}
h1 {{
    color: #2c3e50;
    text-align: center;
    margin-bottom: 5px;
    font-size: 24px;
}}
h2 {{
    color: #34495e;
    margin-top: 40px;
    margin-bottom: 10px;
    padding: 10px 15px;
    background: #2c3e50;
    color: white;
    border-radius: 6px;
    font-size: 16px;
}}
.summary {{
    text-align: center;
    color: #555;
    margin-bottom: 15px;
    font-size: 13px;
}}
.section {{
    margin-bottom: 40px;
}}
.filter-bar {{
    background: #ecf0f1;
    padding: 10px 15px;
    border-radius: 6px;
    margin-bottom: 10px;
    display: flex;
    align-items: center;
    gap: 10px;
    flex-wrap: wrap;
}}
.filter-bar label {{
    font-size: 13px;
    font-weight: bold;
    color: #2c3e50;
}}
.filter-bar select {{
    padding: 5px 10px;
    border-radius: 4px;
    border: 1px solid #bdc3c7;
    font-size: 12px;
    min-width: 200px;
}}
.filter-bar .filter-count {{
    font-size: 11px;
    color: #7f8c8d;
    margin-right: 10px;
}}
.table-container {{
    overflow-x: auto;
    overflow-y: auto;
    max-height: 70vh;
    background: white;
    border-radius: 8px;
    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
    padding: 10px;
}}
table {{
    border-collapse: collapse;
    font-size: 11px;
    white-space: nowrap;
}}
th, td {{
    border: 1px solid #ddd;
    padding: 4px 6px;
    text-align: center;
}}
.m1 th {{
    background: #34495e;
    color: white;
    position: sticky;
    top: 0;
    z-index: 10;
}}
.m1 th.user-header {{
    background: #2c3e50;
    position: sticky;
    right: 0;
    z-index: 20;
    text-align: right;
    min-width: 200px;
}}
.m1 td.user-cell {{
    background: #ecf0f1;
    position: sticky;
    right: 0;
    z-index: 5;
    text-align: right;
    font-weight: bold;
    font-size: 10px;
}}
.m1 th.q-header {{
    writing-mode: vertical-rl;
    text-orientation: mixed;
    max-width: 28px;
    height: 160px;
    font-size: 9px;
    padding: 4px 2px;
}}
.m1 td.answer-ap {{ background: #d4efdf; color: #1e8449; }}
.m1 td.answer-et {{ background: #d6eaf8; color: #1a5276; }}
.m1 td.answer-tp {{ background: #fdebd0; color: #935116; }}
.m1 td.answer-ae {{ background: #fadbd8; color: #922b21; }}
.m1 td.answer-ranking {{ background: #f9f9f9; color: #333; font-size: 9px; }}
.m1 td.answer-pref {{ background: #e8daef; color: #4a235a; font-size: 9px; }}
.m1 td.empty {{ background: #f8f8f8; color: #ccc; }}
.m2 th {{
    background: #34495e;
    color: white;
    position: sticky;
    top: 0;
    z-index: 10;
    font-size: 12px;
}}
.m2 th.q-col {{
    background: #2c3e50;
    position: sticky;
    right: 0;
    z-index: 20;
    text-align: right;
    min-width: 220px;
    max-width: 300px;
    white-space: normal;
    font-size: 11px;
}}
.m2 td.q-cell {{
    background: #f8f9fa;
    position: sticky;
    right: 0;
    z-index: 5;
    text-align: right;
    font-size: 10px;
    max-width: 300px;
    white-space: normal;
}}
.m2 td.stat-cell {{
    font-size: 10px;
    min-width: 80px;
    white-space: normal;
    line-height: 1.4;
}}
.stat-bar {{
    display: flex;
    height: 12px;
    border-radius: 3px;
    overflow: hidden;
    margin-top: 2px;
}}
.bar-segment {{ height: 100%; min-width: 1px; }}
.bar-ap {{ background: #27ae60; }}
.bar-et {{ background: #2980b9; }}
.bar-tp {{ background: #e67e22; }}
.bar-ae {{ background: #c0392b; }}
.bar-d {{ background: #8e44ad; }}
.bar-s {{ background: #16a085; }}
.bar-f {{ background: #f39c12; }}
.bar-other {{ background: #95a5a6; }}
.dominant {{ font-weight: bold; font-size: 10px; }}
.stat-label {{ font-size: 9px; color: #666; }}
.no-data {{ color: #ccc; }}
.phase-row td {{
    background: #7f8c8d;
    color: white;
    font-weight: bold;
    font-size: 11px;
    text-align: center;
}}
.insights-table {{
    width: 100%;
    font-size: 12px;
}}
.insights-table th {{
    background: #2c3e50;
    color: white;
    padding: 8px 10px;
    text-align: right;
}}
.insights-table td {{
    padding: 6px 10px;
    text-align: right;
    white-space: normal;
    vertical-align: top;
}}
.insights-table tr:nth-child(even) {{ background: #f8f9fa; }}
.sig-bar {{
    display: inline-block;
    height: 14px;
    border-radius: 3px;
    min-width: 4px;
    vertical-align: middle;
}}
.sig-high {{ background: #e74c3c; }}
.sig-med {{ background: #f39c12; }}
.sig-low {{ background: #3498db; }}
.sig-none {{ background: #bdc3c7; }}
.code-badge {{
    display: inline-block;
    padding: 2px 6px;
    border-radius: 4px;
    font-size: 10px;
    font-weight: bold;
    margin: 1px;
}}
.code-E {{ background: #d6eaf8; color: #1a5276; }}
.code-A {{ background: #d4efdf; color: #1e8449; }}
.code-T {{ background: #fdebd0; color: #935116; }}
.code-P {{ background: #fadbd8; color: #922b21; }}
.pdn-code {{
    display: inline-block;
    padding: 1px 4px;
    border-radius: 3px;
    font-weight: bold;
    font-size: 10px;
    margin-right: 4px;
}}
.pdn-E {{ background: #d6eaf8; color: #1a5276; }}
.pdn-A {{ background: #d4efdf; color: #1e8449; }}
.pdn-T {{ background: #fdebd0; color: #935116; }}
.pdn-P {{ background: #fadbd8; color: #922b21; }}
.pdn-NA {{ background: #eee; color: #999; }}
.code-counts {{
    display: flex;
    justify-content: center;
    gap: 8px;
    flex-wrap: wrap;
    margin-bottom: 15px;
}}
nav {{
    position: sticky;
    top: 0;
    background: #2c3e50;
    padding: 10px 20px;
    border-radius: 6px;
    margin-bottom: 20px;
    z-index: 100;
    text-align: center;
}}
nav a {{
    color: white;
    text-decoration: none;
    margin: 0 15px;
    font-size: 13px;
    padding: 5px 12px;
    border-radius: 4px;
    transition: background 0.2s;
}}
nav a:hover {{ background: #4a6fa5; }}
tr.hidden {{ display: none; }}
.m2 col.hidden-col {{ visibility: collapse; }}
</style>
</head>
<body>

<h1>PDN - דוח מטריצות משולב (קוד מאבחן)</h1>
<div class="summary">
    סה"כ משתמשים עם קוד מאבחן תקין: {len(users_with_answers)} | שאלות: {len(sorted_q_nums)}<br>
    * הדוח מבוסס על "קוד מאבחן" בלבד (לא קוד מערכת)<br>
    נוצר: {datetime.now().strftime("%d/%m/%Y %H:%M")}
</div>

<nav>
    <a href="#matrix1">מטריצת משתמשים</a>
    <a href="#matrix2">מטריצת קודים</a>
"""

    if insights:
        html += '    <a href="#insights">מובהקות</a>\n'

    html += """</nav>

<!-- ====== SECTION 1: USER ANSWERS MATRIX ====== -->
<div class="section" id="matrix1">
<h2>1. מטריצת תשובות משתמשים</h2>
<div class="filter-bar">
    <label>סנן משתמש:</label>
    <select id="userFilter" onchange="filterUsers()">
        {user_options_html}
    </select>
    <span class="filter-count" id="userCount">מציג {len(users_with_answers)} משתמשים</span>
</div>
<div class="table-container">
<table class="m1" id="userTable">
<thead>
<tr>
<th class="user-header">משתמש (שם / UID / PDN)</th>
"""

    for q_num in sorted_q_nums:
        q = questions[q_num]
        html += f'<th class="q-header" title="Q{q_num}: {q["text"]} [{q["codes"]}]">Q{q_num}</th>\n'

    html += "</tr>\n</thead>\n<tbody>\n"

    for user in users:
        email = user["email"]
        answers = all_answers.get(email, {})
        if not answers:
            continue
        uid = user["uid"]
        pdn = user["pdn_code"] or "NA"
        pdn_class = f"pdn-{pdn[0]}" if pdn and pdn[0] in "EATP" else "pdn-NA"
        first_name, last_name = user_names.get(email, ("", ""))
        display_name = f"{first_name} {last_name}".strip()

        html += f'<tr data-uid="{uid}" data-pdn="{pdn}">'
        html += f'<td class="user-cell"><span class="pdn-code {pdn_class}">{pdn}</span> {uid}'
        if display_name:
            html += f' <span style="color:#555;font-weight:normal;">{display_name}</span>'
        html += '</td>\n'

        for q_num in sorted_q_nums:
            answer = answers.get(str(q_num)) or answers.get(q_num)
            if answer is None or (isinstance(answer, dict) and "selected_option_code" not in answer and "ranking" not in answer):
                html += '<td class="empty">-</td>\n'
                continue
            formatted = format_answer_display(answer, q_num)
            if not formatted:
                html += '<td class="empty">-</td>\n'
                continue
            cell_class = ""
            if formatted in ("AP", "A"):
                cell_class = "answer-ap"
            elif formatted in ("ET", "TE", "E"):
                cell_class = "answer-et"
            elif formatted in ("TP", "T"):
                cell_class = "answer-tp"
            elif formatted in ("AE", "P"):
                cell_class = "answer-ae"
            elif ":" in formatted:
                cell_class = "answer-ranking"
            else:
                cell_class = "answer-pref"
            html += f'<td class="{cell_class}" title="{formatted}">{formatted}</td>\n'

        html += "</tr>\n"

    html += """</tbody></table></div></div>

<!-- ====== SECTION 2: PDN CODE STATS MATRIX ====== -->
<div class="section" id="matrix2">
<h2>2. מטריצת סטטיסטיקה לפי קוד PDN</h2>
<div class="filter-bar">
    <label>סנן קוד PDN:</label>
    <select id="pdnFilter" onchange="filterPdnCodes()">
        <option value="all" selected>הכל</option>
"""

    for code in PDN_12_CODES:
        count = code_counts.get(code, 0)
        html += f'<option value="{code}">{code} ({count} משתמשים)</option>\n'

    html += """    </select>
    <span class="filter-count" id="pdnCount">מציג 12 קודים</span>
</div>
<div class="code-counts">
"""

    for code in PDN_12_CODES:
        css_class = f"code-{code[0]}"
        count = code_counts.get(code, 0)
        html += f'<span class="code-badge {css_class}">{code}: {count}</span>\n'

    html += """</div>
<div class="table-container">
<table class="m2" id="pdnTable">
<thead>
<tr>
<th class="q-col">שאלה</th>
"""

    for i, code in enumerate(PDN_12_CODES):
        html += f'<th class="pdn-col" data-code="{code}">{code}</th>\n'

    html += "</tr>\n</thead>\n<tbody>\n"

    current_phase = None
    for q_num in sorted_q_nums:
        q = questions[q_num]
        if q["phase"] != current_phase:
            current_phase = q["phase"]
            phase_label = {
                "PartA": "חלק א - בחירה בינארית (AP/ET, TP/AE)",
                "PartB": "חלק ב - דירוג (D/S/F)",
                "PartC": "חלק ג - סולם (A/T, P/E)",
                "PartD": "חלק ד - ילדות/בגרות (TP/AE, AP/TE)",
                "PartE": "חלק ה - דירוג 4 (E/P/T/A)",
            }.get(current_phase, current_phase)
            html += f'<tr class="phase-row"><td colspan="{len(PDN_12_CODES) + 1}">{phase_label}</td></tr>\n'

        short_text = q["text"][:50] + ("..." if len(q["text"]) > 50 else "")
        html += f'<tr><td class="q-cell"><strong>Q{q_num}</strong> [{q["codes"]}]<br>{short_text}</td>\n'

        for pdn_code in PDN_12_CODES:
            stat = stats[q_num].get(pdn_code, {"counts": {}, "total": 0})
            total = stat["total"]
            counts = stat["counts"]

            if total == 0:
                html += f'<td class="stat-cell no-data pdn-col" data-code="{pdn_code}">-</td>\n'
                continue

            sorted_answers = sorted(counts.items(), key=lambda x: -x[1])
            dominant = sorted_answers[0]

            cell_parts = []
            for ans, cnt in sorted_answers[:2]:
                pct = round(cnt / total * 100)
                is_dom = (ans == dominant[0])
                cls = "dominant" if is_dom else "stat-label"
                ans_text = answer_texts.get(q_num, {}).get(ans, "")
                if is_dom and ans_text:
                    display = f"{ans}:{pct}% ({ans_text[:15]})"
                else:
                    display = f"{ans}:{pct}%"
                cell_parts.append(f'<span class="{cls}">{display}</span>')

            color_map = {
                "AP": "bar-ap", "ET": "bar-et", "TE": "bar-et", "TP": "bar-tp", "AE": "bar-ae",
                "D": "bar-d", "S": "bar-s", "F": "bar-f",
                "A": "bar-ap", "T": "bar-tp", "E": "bar-et", "P": "bar-ae",
            }
            bar_html = '<div class="stat-bar">'
            for ans, cnt in sorted_answers:
                pct = cnt / total * 100
                bar_class = color_map.get(ans, "bar-other")
                bar_html += f'<div class="bar-segment {bar_class}" style="width:{pct}%"></div>'
            bar_html += '</div>'

            html += f'<td class="stat-cell pdn-col" data-code="{pdn_code}">{" ".join(cell_parts)}{bar_html}</td>\n'

        html += "</tr>\n"

    html += """</tbody></table></div></div>

"""

    # Only include insights section if there are insights to show
    if insights:
        html += """
<!-- ====== SECTION 3: INSIGHTS ====== -->
<div class="section" id="insights">
<h2>3. ניתוח מובהקות - עד כמה כל שאלה מבדלת בין קודים</h2>
<p style="font-size:12px;color:#666;margin-bottom:10px;">
ציון מובהקות גבוה = השאלה מבדלת היטב בין קודי PDN שונים (קודים שונים עונים תשובות שונות).<br>
ציון נמוך = כל הקודים עונים בצורה דומה (השאלה לא מבחינה בין הקודים).
</p>
<div class="table-container" style="max-height:none;">
<table class="insights-table">
<thead>
<tr>
<th style="width:40px;">#</th>
<th style="width:60px;">שאלה</th>
<th style="width:200px;">תיאור</th>
<th style="width:60px;">קודים</th>
<th style="width:80px;">מובהקות</th>
<th style="width:100px;">ציון</th>
<th style="width:120px;">הערכה</th>
<th>תשובה דומיננטית לפי קוד</th>
</tr>
</thead>
<tbody>
"""

        for i, ins in enumerate(insights, 1):
            sig = ins["significance"]
            if sig >= 70:
                sig_class = "sig-high"
                sig_color = "#e74c3c"
            elif sig >= 40:
                sig_class = "sig-med"
                sig_color = "#f39c12"
            elif sig >= 20:
                sig_class = "sig-low"
                sig_color = "#3498db"
            else:
                sig_class = "sig-none"
                sig_color = "#bdc3c7"

            bar_width = max(4, sig)

            dom_parts = []
            for code, val in ins["dominant_by_code"].items():
                css = f"code-{code[0]}"
                dom_parts.append(f'<span class="code-badge {css}">{code}:{val}</span>')
            dom_html = " ".join(dom_parts[:6])

            html += f"""<tr>
<td>{i}</td>
<td><strong>Q{ins["q_num"]}</strong></td>
<td>{ins["text"][:60]}</td>
<td>{ins["codes"]}</td>
<td><span class="sig-bar {sig_class}" style="width:{bar_width}px;"></span> {sig}</td>
<td style="color:{sig_color};font-weight:bold;">{ins["insight"][:20]}</td>
<td>{ins["insight"]}</td>
<td style="font-size:10px;">{dom_html}</td>
</tr>\n"""

        html += """</tbody></table></div></div>
"""

    html += """
<script>
function filterUsers() {
    const filter = document.getElementById('userFilter').value;
    const rows = document.querySelectorAll('#userTable tbody tr');
    let visible = 0;
    rows.forEach(row => {
        if (filter === 'all') {
            row.classList.remove('hidden');
            visible++;
        } else {
            const uid = row.getAttribute('data-uid');
            if (uid === filter) {
                row.classList.remove('hidden');
                visible++;
            } else {
                row.classList.add('hidden');
            }
        }
    });
    document.getElementById('userCount').textContent = 'מציג ' + visible + ' משתמשים';
}

function filterPdnCodes() {
    const filter = document.getElementById('pdnFilter').value;
    const headerCells = document.querySelectorAll('#pdnTable thead th.pdn-col');
    const bodyCells = document.querySelectorAll('#pdnTable tbody td.pdn-col');
    let visible = 0;

    headerCells.forEach(th => {
        const code = th.getAttribute('data-code');
        if (filter === 'all' || code === filter) {
            th.style.display = '';
            visible++;
        } else {
            th.style.display = 'none';
        }
    });

    bodyCells.forEach(td => {
        const code = td.getAttribute('data-code');
        if (filter === 'all' || code === filter) {
            td.style.display = '';
        } else {
            td.style.display = 'none';
        }
    });

    document.getElementById('pdnCount').textContent = 'מציג ' + (filter === 'all' ? '12' : '1') + ' קודים';
}
</script>
</body>
</html>"""
    return html


# --- Excel Generation ---

def generate_excel(users, questions, all_answers, stats, users_by_code, answer_texts, user_names):
    """Generate an Excel workbook with 3 sheets matching the HTML report."""
    if not OPENPYXL_AVAILABLE:
        print("ERROR: openpyxl is not installed. Run: pip install openpyxl")
        return None

    wb = openpyxl.Workbook()

    # --- Color helpers ---
    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _font(bold=False, color="000000", size=10):
        return Font(bold=bold, color=color, size=size)

    def _border():
        thin = Side(style="thin", color="CCCCCC")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    CODE_FILLS = {
        "E": _fill("D6EAF8"),
        "A": _fill("D4EFDF"),
        "T": _fill("FDEBD0"),
        "P": _fill("FADBD8"),
    }
    ANSWER_FILLS = {
        "AP": _fill("D4EFDF"),
        "ET": _fill("D6EAF8"),
        "TE": _fill("D6EAF8"),  # same as ET
        "AE": _fill("FADBD8"),
        "TP": _fill("FDEBD0"),
        "A":  _fill("D4EFDF"),
        "E":  _fill("D6EAF8"),
        "P":  _fill("FADBD8"),
        "T":  _fill("FDEBD0"),
    }
    HEADER_FILL = _fill("2C3E50")
    HEADER_FONT = _font(bold=True, color="FFFFFF", size=10)
    PHASE_FILL  = _fill("7F8C8D")
    PHASE_FONT  = _font(bold=True, color="FFFFFF", size=10)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=True)

    def sort_key(q_num):
        try:
            return int(q_num)
        except ValueError:
            return 999

    sorted_q_nums = sorted(questions.keys(), key=sort_key)
    users_with_answers = [u for u in users if u["email"] in all_answers and all_answers[u["email"]]]

    # Legacy code aliases for questions whose option codes were changed after deployment
    # Q14 was re-coded from AP/ET -> TP/AE; text is identical, add old codes as aliases
    _legacy_code_text = {
        '14': {'AP': 'נוטה להסכים', 'ET': 'נוטה להתדיין'},
    }

    def _get_option_text(q_num, code):
        """Lookup option text, falling back to legacy aliases for re-coded questions."""
        opts = questions.get(q_num, {}).get("options", [])
        text = next((o.get("text", "") for o in opts if o.get("code") == code), "")
        if not text:
            text = _legacy_code_text.get(q_num, {}).get(code, "")
        return text[:20]

    # ================================================================
    # Sheet 1 - User Answers Matrix
    # ================================================================
    ws1 = wb.active
    ws1.title = "תשובות משתמשים"
    ws1.sheet_view.rightToLeft = True

    # Header row
    ws1.cell(1, 1, "משתמש (PDN | UID | שם)").font = HEADER_FONT
    ws1.cell(1, 1).fill = HEADER_FILL
    ws1.cell(1, 1).alignment = RIGHT
    ws1.column_dimensions["A"].width = 28

    for col_idx, q_num in enumerate(sorted_q_nums, start=2):
        c = ws1.cell(1, col_idx, f"Q{q_num}")
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = _border()
        ws1.column_dimensions[get_column_letter(col_idx)].width = 7

    ws1.row_dimensions[1].height = 30
    ws1.freeze_panes = "B2"

    # Data rows
    for row_idx, user in enumerate(users_with_answers, start=2):
        email  = user["email"]
        answers = all_answers.get(email, {})
        uid    = user["uid"]
        pdn    = user["pdn_code"] or "NA"
        first, last = user_names.get(email, ("", ""))
        display = f"{first} {last}".strip()
        label  = f"{pdn} | {uid}" + (f" | {display}" if display else "")

        user_cell = ws1.cell(row_idx, 1, label)
        user_cell.alignment = RIGHT
        user_cell.border = _border()
        if pdn and pdn[0] in CODE_FILLS:
            user_cell.fill = CODE_FILLS[pdn[0]]

        for col_idx, q_num in enumerate(sorted_q_nums, start=2):
            answer = answers.get(str(q_num)) or answers.get(q_num)
            val = format_answer_code(answer, q_num) if answer else None
            if val and answer and 'ranking' in answer and _is_rank_order(q_num):
                # PartB/E: show full ranking e.g. "S:1 F:2 D:3\n(פועל בקצב שלי)"
                ranking = answer['ranking']
                sorted_r = sorted(ranking.items(), key=lambda x: x[1])
                rank_str = " ".join(f"{k}:{v}" for k, v in sorted_r)
                top_text = _get_option_text(q_num, val)
                display = f"{rank_str}\n({top_text})" if top_text else rank_str
            elif val and answer and 'ranking' in answer:
                # PartC/D scale: show "TP:10 AE:2\n(מופנם מאוד)"
                ranking = answer['ranking']
                sorted_r = sorted(ranking.items(), key=lambda x: -x[1])
                scale_str = " ".join(f"{k}:{v}" for k, v in sorted_r)
                top_text = _get_option_text(q_num, val)
                display = f"{scale_str}\n({top_text})" if top_text else scale_str
            elif val:
                top_text = _get_option_text(q_num, val)
                display = f"{val} ({top_text})" if top_text else val
            else:
                display = ""
            c = ws1.cell(row_idx, col_idx, display)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _border()
            if val and val in ANSWER_FILLS:
                c.fill = ANSWER_FILLS[val]

    # ================================================================
    # Sheet 2 - PDN Code Stats Matrix
    # ================================================================
    ws2 = wb.create_sheet("סטטיסטיקה לפי קוד")
    ws2.sheet_view.rightToLeft = True

    # Header
    ws2.cell(1, 1, "שאלה").font = HEADER_FONT
    ws2.cell(1, 1).fill = HEADER_FILL
    ws2.cell(1, 1).alignment = RIGHT
    ws2.column_dimensions["A"].width = 40

    for col_idx, code in enumerate(PDN_12_CODES, start=2):
        count = len(users_by_code.get(code, []))
        c = ws2.cell(1, col_idx, f"{code}\n({count})")
        c.font = HEADER_FONT
        c.fill = CODE_FILLS.get(code[0], HEADER_FILL) if code[0] in CODE_FILLS else HEADER_FILL
        c.font = _font(bold=True, color="000000" if code[0] in CODE_FILLS else "FFFFFF", size=10)
        c.alignment = CENTER
        c.border = _border()
        ws2.column_dimensions[get_column_letter(col_idx)].width = 16

    ws2.row_dimensions[1].height = 36
    ws2.freeze_panes = "B2"

    current_phase = None
    row_idx = 2
    for q_num in sorted_q_nums:
        q = questions[q_num]

        # Phase separator row
        if q["phase"] != current_phase:
            current_phase = q["phase"]
            phase_label = {
                "PartA": "חלק א - בחירה בינארית",
                "PartB": "חלק ב - דירוג",
                "PartC": "חלק ג - סולם",
                "PartD": "חלק ד - ילדות/בגרות",
                "PartE": "חלק ה - דירוג 4",
            }.get(current_phase, current_phase)
            for col in range(1, len(PDN_12_CODES) + 2):
                c = ws2.cell(row_idx, col, phase_label if col == 1 else "")
                c.fill = PHASE_FILL
                c.font = PHASE_FONT
                c.alignment = CENTER
                c.border = _border()
            ws2.row_dimensions[row_idx].height = 18
            row_idx += 1

        # Question label
        short_text = q["text"][:55] + ("..." if len(q["text"]) > 55 else "")
        q_label = f"Q{q_num} [{q['codes']}]\n{short_text}"
        lc = ws2.cell(row_idx, 1, q_label)
        lc.alignment = RIGHT
        lc.border = _border()
        ws2.row_dimensions[row_idx].height = 36

        # Stats per PDN code
        for col_idx, pdn_code in enumerate(PDN_12_CODES, start=2):
            s = stats[q_num].get(pdn_code, {"counts": {}, "total": 0})
            total = s["total"]
            if total == 0:
                c = ws2.cell(row_idx, col_idx, "-")
                c.alignment = CENTER
                c.border = _border()
                continue

            sorted_answers = sorted(s["counts"].items(), key=lambda x: -x[1])
            lines = []
            for ans, cnt in sorted_answers[:3]:
                pct = round(cnt / total * 100)
                txt = answer_texts.get(q_num, {}).get(ans, "")[:18]
                lines.append(f"{ans}: {pct}%" + (f" ({txt})" if txt else ""))
            cell_val = "\n".join(lines)

            c = ws2.cell(row_idx, col_idx, cell_val)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _border()
            # Color by dominant answer
            dom_ans = sorted_answers[0][0]
            if dom_ans in ANSWER_FILLS:
                c.fill = ANSWER_FILLS[dom_ans]

        row_idx += 1

    # ================================================================
    # Sheet 3 - Summary (user count per code)
    # ================================================================
    ws3 = wb.create_sheet("סיכום קודים")
    ws3.sheet_view.rightToLeft = True

    ws3.cell(1, 1, "קוד PDN").font = _font(bold=True, size=11)
    ws3.cell(1, 2, "מספר מאובחנים").font = _font(bold=True, size=11)
    ws3.cell(1, 1).fill = HEADER_FILL
    ws3.cell(1, 2).fill = HEADER_FILL
    ws3.cell(1, 1).font = HEADER_FONT
    ws3.cell(1, 2).font = HEADER_FONT
    ws3.cell(1, 1).alignment = CENTER
    ws3.cell(1, 2).alignment = CENTER
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 20

    for row_idx, code in enumerate(PDN_12_CODES, start=2):
        count = len(users_by_code.get(code, []))
        c1 = ws3.cell(row_idx, 1, code)
        c2 = ws3.cell(row_idx, 2, count)
        c1.alignment = CENTER
        c2.alignment = CENTER
        c1.border = _border()
        c2.border = _border()
        if code[0] in CODE_FILLS:
            c1.fill = CODE_FILLS[code[0]]
            c2.fill = CODE_FILLS[code[0]]

    ws3.cell(len(PDN_12_CODES) + 2, 1, "סה\"כ").font = _font(bold=True)
    ws3.cell(len(PDN_12_CODES) + 2, 2, len(users_with_answers)).font = _font(bold=True)

    return wb


# --- Main ---

def main():
    # Parse arguments
    include_significance = "--significance" in sys.argv or "--מובהקות" in sys.argv

    # --format html | excel | both  (default: html)
    fmt = "html"
    for arg in sys.argv[1:]:
        if arg.startswith("--format="):
            fmt = arg.split("=", 1)[1].lower()
        elif arg in ("--format", "-f"):
            idx = sys.argv.index(arg)
            if idx + 1 < len(sys.argv):
                fmt = sys.argv[idx + 1].lower()

    if fmt not in ("html", "excel", "both"):
        print(f"Unknown format '{fmt}'. Valid options: html, excel, both. Defaulting to html.")
        fmt = "html"

    if fmt in ("excel", "both") and not OPENPYXL_AVAILABLE:
        print("ERROR: openpyxl is required for Excel output.")
        print("Install it with:  pip install openpyxl")
        sys.exit(1)

    print("Loading questions...")
    questions = load_questions()
    print(f"  {len(questions)} questions")

    print("Loading users...")
    users = load_users()
    print(f"  {len(users)} users")

    print("Computing statistics...")
    stats, users_by_code, all_answers, answer_texts, user_names = compute_statistics(users, questions)
    answered = sum(1 for v in all_answers.values() if v)
    print(f"  {answered} users with answers")

    insights = []
    if include_significance:
        print("Computing significance (מובהקות)...")
        insights = compute_significance(stats, questions, users_by_code)
        high_sig = sum(1 for i in insights if i["significance"] >= 70)
        print(f"  {high_sig} questions with high significance")
    else:
        print("  Skipping significance section (use --significance to include)")

    output_dir = PROJECT_ROOT / "docs"
    output_dir.mkdir(parents=True, exist_ok=True)
    saved = []

    # HTML output
    if fmt in ("html", "both"):
        print("Generating HTML report...")
        html = generate_html(users, questions, all_answers, stats, users_by_code, answer_texts, insights, user_names)
        html_path = output_dir / "combined_matrix_report.html"
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html)
        print(f"  HTML saved to: {html_path}")
        saved.append(str(html_path))

    # Excel output
    if fmt in ("excel", "both"):
        print("Generating Excel report...")
        wb = generate_excel(users, questions, all_answers, stats, users_by_code, answer_texts, user_names)
        if wb:
            xlsx_path = output_dir / "combined_matrix_report.xlsx"
            wb.save(xlsx_path)
            print(f"  Excel saved to: {xlsx_path}")
            saved.append(str(xlsx_path))

    print("\nDone! Opening report(s)...")
    for path in saved:
        os.system(f"open '{path}'")

    return saved[0] if saved else None


if __name__ == "__main__":
    main()
