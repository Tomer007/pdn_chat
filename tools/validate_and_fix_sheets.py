#!/usr/bin/env python3
"""
Validate and fix the Excel statistics sheets.

Issues found:
1. Q38-Q56: percentages rounded with // causing 66% instead of 67%, 33% instead of 33% - acceptable
2. Q57-Q61: Only shows first choice (position 1). 
   Requirement asks: for each PDN code + each Q57-Q61, show how many users
   put each trait (T/A/E/P) in position 1, 2, 3, 4
3. Need to use full ranking data from JSON files (not just Excel Q57 column which only stores pos1)
"""

import json, glob, os
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ===========================
# LOAD DATA
# ===========================

# Load PDN code from users.json (email -> pdn_code)
with open('/Users/tomer.gur/dev-tools/pdn_chat/app/data/users.json', 'r', encoding='utf-8') as f:
    users_db = json.load(f)
email_to_pdn = {email.lower(): data['pdn_code'].upper() for email, data in users_db.items()}

# Also load PDN codes from the Excel for the diagnosed users
wb = openpyxl.load_workbook('/Users/tomer.gur/dev-tools/pdn_chat/docs/combined_matrix_report.xlsx')
ws_answers = wb['תשובות משתמשים']
header = [cell.value for cell in ws_answers[1]]

# Extract email->pdn from Excel user string (P10 | UID | Name)
# We need to also load from metadata in json files
excel_users_pdn = {}
for row in ws_answers.iter_rows(min_row=2, values_only=True):
    user_str = row[0]
    if not user_str:
        continue
    parts = [p.strip() for p in str(user_str).split('|')]
    if len(parts) >= 2:
        pdn_code = parts[0].strip()
        uid = parts[1].replace('UID', '').strip().lower()
        excel_users_pdn[uid] = pdn_code

PDN_CODES = ['E1', 'E5', 'E9', 'A3', 'A7', 'A11', 'T4', 'T8', 'T12', 'P2', 'P6', 'P10']

# Load all JSON answer files
json_files = glob.glob('/Users/tomer.gur/dev-tools/pdn_chat/saved_results/**/*_answers.json', recursive=True)

# Build dataset: pdn_code -> Q38-Q56 answers and Q57-Q61 rankings
data_by_code = defaultdict(lambda: {
    'q38_q56': defaultdict(list),
    'q57_q61_rankings': defaultdict(lambda: defaultdict(lambda: defaultdict(int))),  # q -> pos -> trait -> count
    'q57_q61_first': defaultdict(list),
    'count': 0
})

matched_count = 0
for jf in json_files:
    with open(jf, 'r', encoding='utf-8') as f:
        try:
            data = json.load(f)
        except:
            continue
    
    meta = data.get('metadata', {})
    email = meta.get('email', '').lower().strip()
    
    # Find PDN code
    pdn_code = email_to_pdn.get(email)
    if not pdn_code:
        continue
    
    pdn_code = pdn_code.upper()
    if pdn_code not in PDN_CODES:
        continue
    
    data_by_code[pdn_code]['count'] += 1
    matched_count += 1
    
    # Q38-Q56 binary answers
    for q_num in range(38, 57):
        q_str = str(q_num)
        if q_str in data and isinstance(data[q_str], dict):
            ans = data[q_str].get('selected_option_code')
            if ans:
                data_by_code[pdn_code]['q38_q56'][q_num].append(ans)
    
    # Q57-Q61 full rankings
    for q_num in range(57, 62):
        q_str = str(q_num)
        if q_str not in data:
            continue
        q_data = data[q_str]
        if isinstance(q_data, dict):
            ranking = q_data.get('ranking', {})
            if ranking and isinstance(ranking, dict):
                for trait, pos in ranking.items():
                    if trait in ['T', 'A', 'E', 'P'] and isinstance(pos, int):
                        data_by_code[pdn_code]['q57_q61_rankings'][q_num][pos][trait] += 1
            # Also store first choice
            if ranking:
                first = min(ranking.items(), key=lambda x: x[1] if isinstance(x[1], int) else 99)
                if first[0] in ['T', 'A', 'E', 'P']:
                    data_by_code[pdn_code]['q57_q61_first'][q_num].append(first[0])

print(f'Matched {matched_count} users')
for code in PDN_CODES:
    n = data_by_code[code]['count']
    if n > 0:
        print(f'  {code}: N={n}')

# ===========================
# STYLES
# ===========================
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
subheader_fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
subheader_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def pct(count, total):
    if total == 0:
        return '-'
    return f'{round(count * 100 / total)}%'

# ===========================
# Q38-Q56 SHEET (rebuild)
# ===========================
Q38_56_QUESTIONS = {
    38: "בדיקה לעומק לפני סיכון",
    39: "קושי עם חוסר ודאות",
    40: "קבלת החלטות במהירות",
    41: "קושי לקבל הנחיות",
    42: "נוחות להוביל אחרים",
    43: "מוחצנות/מופנמות - ילדות",
    44: "מוחצנות/מופנמות - בגרות",
    45: "הובלה/הצטרפות - ילדות",
    46: "הובלה/הצטרפות - בגרות",
    47: "דברנות/שתקנות - ילדות",
    48: "דברנות/שתקנות - בגרות",
    49: "עימותים/ריצוי - ילדות",
    50: "עימותים/ריצוי - בגרות",
    51: "עמידה בזמנים - ילדות",
    52: "עמידה בזמנים - בגרות",
    53: "סדר/בלגן - ילדות",
    54: "סדר/בלגן - בגרות",
    55: "מאופקות/נועזות - ילדות",
    56: "מאופקות/נועזות - בגרות",
}

if 'Q38-Q56 סטטיסטיקה' in wb.sheetnames:
    del wb['Q38-Q56 סטטיסטיקה']
ws38 = wb.create_sheet('Q38-Q56 סטטיסטיקה')

# Header row
row1 = ['שאלה (Q38-Q56)']
for code in PDN_CODES:
    n = data_by_code[code]['count']
    row1.append(f'{code}\n(N={n})')

for col, val in enumerate(row1, 1):
    c = ws38.cell(row=1, column=col, value=val)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin_border

# Data rows
for row_idx, q_num in enumerate(range(38, 57), 2):
    q_text = Q38_56_QUESTIONS[q_num]
    c = ws38.cell(row=row_idx, column=1, value=f'Q{q_num} - {q_text}')
    c.border = thin_border
    c.alignment = Alignment(wrap_text=True, vertical='center')
    
    for col_idx, code in enumerate(PDN_CODES, 2):
        answers = data_by_code[code]['q38_q56'].get(q_num, [])
        total = len(answers)
        if total == 0:
            c = ws38.cell(row=row_idx, column=col_idx, value='-')
        else:
            counts = defaultdict(int)
            for a in answers:
                counts[a] += 1
            lines = []
            for trait, cnt in sorted(counts.items(), key=lambda x: -x[1]):
                lines.append(f'{trait}: {pct(cnt, total)} ({cnt}/{total})')
            c = ws38.cell(row=row_idx, column=col_idx, value='\n'.join(lines))
        c.border = thin_border
        c.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

ws38.column_dimensions['A'].width = 30
ws38.row_dimensions[1].height = 40
for q in range(38, 57):
    ws38.row_dimensions[q - 38 + 2].height = 35
for col in range(2, len(PDN_CODES) + 2):
    ws38.column_dimensions[get_column_letter(col)].width = 14

print('Q38-Q56 sheet created')

# ===========================
# Q57-Q61 SHEET (rebuild with full ranking positions)
# ===========================
Q57_61_QUESTIONS = {
    57: "תכונות: אסרטיבי/אכפתי/שיטתי/שופע רעיונות",
    58: "תכונות: מוביל/תומך/זהיר/רעיוניסט",
    59: "תכונות: ממוקד מטרה/נותן/שיטתי/אופטימי",
    60: "מה רוצה לקבל מהאבחון",
    61: "מה חשוב שהאבחון יספק",
}

if 'Q57-Q61 סטטיסטיקה' in wb.sheetnames:
    del wb['Q57-Q61 סטטיסטיקה']
ws57 = wb.create_sheet('Q57-Q61 סטטיסטיקה')

# For Q57-Q61 we show for each PDN code:
# % of users who put each trait in position 1 (first choice)
# Then N users per position per trait

row_num = 1
# Title
for col, val in enumerate(['שאלה (Q57-Q61)'] + [f'{code}\n(N={data_by_code[code]["count"]})' for code in PDN_CODES], 1):
    c = ws57.cell(row=row_num, column=col, value=val)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin_border

for q_num in range(57, 62):
    row_num += 1
    q_text = Q57_61_QUESTIONS[q_num]
    c = ws57.cell(row=row_num, column=1, value=f'Q{q_num}\n{q_text}')
    c.border = thin_border
    c.fill = subheader_fill
    c.font = subheader_font
    c.alignment = Alignment(wrap_text=True, vertical='center')
    
    for col_idx, code in enumerate(PDN_CODES, 2):
        rankings = data_by_code[code]['q57_q61_rankings'].get(q_num, {})
        first_choices = data_by_code[code]['q57_q61_first'].get(q_num, [])
        total = len(first_choices)
        
        if total == 0:
            c = ws57.cell(row=row_num, column=col_idx, value='-')
        else:
            lines = []
            # First choice % for each trait
            first_counts = defaultdict(int)
            for t in first_choices:
                first_counts[t] += 1
            
            for trait in ['E', 'P', 'T', 'A']:
                cnt = first_counts.get(trait, 0)
                if cnt > 0:
                    lines.append(f'{trait}: {pct(cnt, total)}')
            
            # Position distribution: how many put each trait in pos1/2/3/4
            if rankings:
                lines.append('---')
                for pos in [1, 2, 3, 4]:
                    pos_data = rankings.get(pos, {})
                    if pos_data:
                        pos_str = f'מקום {pos}: ' + ' '.join([f'{t}={c}' for t, c in sorted(pos_data.items())])
                        lines.append(pos_str)
            
            c = ws57.cell(row=row_num, column=col_idx, value='\n'.join(lines))
        c.border = thin_border
        c.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

ws57.column_dimensions['A'].width = 35
ws57.row_dimensions[1].height = 40
for q in range(57, 62):
    ws57.row_dimensions[q - 57 + 2].height = 80
for col in range(2, len(PDN_CODES) + 2):
    ws57.column_dimensions[get_column_letter(col)].width = 16

print('Q57-Q61 sheet created')

# Save
wb.save('/Users/tomer.gur/dev-tools/pdn_chat/docs/combined_matrix_report.xlsx')
print('Saved.')

# ===========================
# VALIDATION REPORT
# ===========================
print('\n=== VALIDATION REPORT ===')
print()
print('REQ 1: גיליון תשובות מאורגנים (Name | UID | Verified PDN Code)')
ws_org = wb['גיליון תשובות מאורגנים']
headers = [cell.value for cell in ws_org[1]]
print(f'  Headers: {headers}')
print(f'  Rows: {ws_org.max_row - 1}')
print(f'  PASS: {headers == ["Name", "UID", "Verified PDN Code"]}')

print()
print('REQ 2: Q38-Q56 statistics by PDN code')
ws38v = wb['Q38-Q56 סטטיסטיקה']
print(f'  Rows (questions): {ws38v.max_row - 1} (expected 19)')
print(f'  Cols: {ws38v.max_column} (expected 13 = 1 label + 12 codes)')
# Sample check E1 Q38
e1_q38 = ws38v.cell(row=2, column=2).value
print(f'  E1/Q38 value: {e1_q38}')
print(f'  PASS: {ws38v.max_row - 1 == 19 and ws38v.max_column == 13}')

print()
print('REQ 3: Q57-Q61 statistics by PDN code (first choice + position distribution)')
ws57v = wb['Q57-Q61 סטטיסטיקה']
print(f'  Rows (questions): {ws57v.max_row - 1} (expected 5)')
print(f'  Cols: {ws57v.max_column} (expected 13)')
# Sample check
e5_q57 = ws57v.cell(row=2, column=3).value
print(f'  E5/Q57 value: {e5_q57}')
print(f'  PASS: {ws57v.max_row - 1 == 5 and ws57v.max_column == 13}')
