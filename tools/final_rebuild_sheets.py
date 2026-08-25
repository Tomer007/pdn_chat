#!/usr/bin/env python3
"""
Rebuild PDN statistics Excel sheets from production data.
Loads data from /tmp/pdn_analysis_full.json (fetched from production API).
Includes: LLR type analysis, comparison table, per-code accuracy summary.

Features used for LLR model:
- Q1-Q26: selected_option_code (AP / ET / AE / TP)
- Q27-Q37: ranking rank per letter (D/F/S: 1-3)
- Q38-Q42: ranking score per letter (A/T or E/P: 0-12)
- Q43-Q56: ranking score per pair (AE/TP or AP/TE: 0-12)
- Q57-Q61: ranking rank per letter (A/E/P/T: 1-4)
"""

import json, math
from collections import defaultdict
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── paths ──────────────────────────────────────────────────────────────────────
BASE       = Path(__file__).resolve().parent.parent
EXCEL      = BASE / "statistics" / "pdn_statistics.xlsx"
DATA_JSON  = Path("/tmp/pdn_analysis_full.json")

# ── constants ──────────────────────────────────────────────────────────────────
PDN_CODES = ["E1", "E5", "E9", "A3", "A7", "A11", "T4", "T8", "T12", "P2", "P6", "P10"]

# ── styles ─────────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="2F4F8F")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL      = PatternFill("solid", fgColor="EEF2FF")
MATCH_FILL    = PatternFill("solid", fgColor="C6EFCE")
MISMATCH_FILL = PatternFill("solid", fgColor="FFC7CE")
SECTION_FILL  = PatternFill("solid", fgColor="D9E1F2")
EMPTY_FILL    = PatternFill()

def _hdr(ws, row, col, text, width=None):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def _cell(ws, row, col, value, fill=None, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    if fill:
        c.fill = fill
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    return c

# ── feature extraction ─────────────────────────────────────────────────────────
def extract_features(answers: dict) -> dict:
    """
    Convert raw answer dict to flat feature dict for LLR model.
    Returns: {feature_name: value_string}
    """
    feats = {}
    for qnum_str, ans in answers.items():
        qnum = int(qnum_str)

        if "selected_option_code" in ans:
            # Q1-Q26: binary choice code
            feats[f"q{qnum}_code"] = ans["selected_option_code"]

        elif "ranking" in ans:
            ranking = ans["ranking"]
            if qnum <= 61:
                # Store each letter's value as a separate feature
                for letter, val in ranking.items():
                    feats[f"q{qnum}_{letter}"] = str(val)

    return feats

# ── LLR model ─────────────────────────────────────────────────────────────────
def train_llr_model(users: list, answers: dict) -> dict:
    """
    Train LLR model from labeled users.
    Returns: {code: {feature: {value: log_likelihood_ratio}}}
    """
    labeled = [(u["email"], u["pdn_code"]) for u in users if u["pdn_code"] in PDN_CODES]

    if len(labeled) < 10:
        return {}

    feat_count = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    code_total = defaultdict(int)

    for email, code in labeled:
        code_total[code] += 1
        user_ans = answers.get(email, {})
        feats = extract_features(user_ans)
        for feat, val in feats.items():
            feat_count[feat][val][code] += 1

    total = sum(code_total.values())

    model = {}
    for code in PDN_CODES:
        model[code] = {}
        n_code = code_total.get(code, 0)
        n_not  = total - n_code
        if n_code < 2:
            continue
        for feat, val_dict in feat_count.items():
            model[code][feat] = {}
            for val, code_counts in val_dict.items():
                c_in  = code_counts.get(code, 0)
                c_out = sum(v for k, v in code_counts.items() if k != code)
                p_in  = (c_in  + 0.5) / (n_code + 1)
                p_out = (c_out + 0.5) / (n_not  + 1)
                model[code][feat][val] = math.log(p_in / p_out) if p_out > 0 else 0.0

    return model

def compute_llr_scores(user_answers: dict, model: dict) -> dict:
    """Compute LLR score for each PDN code. Returns {code: score}."""
    feats = extract_features(user_answers)
    scores = {}
    for code in PDN_CODES:
        score = 0.0
        code_model = model.get(code, {})
        for feat, val in feats.items():
            if feat in code_model:
                score += code_model[feat].get(val, 0.0)
        scores[code] = score
    return scores

def get_suggestion(email: str, answers: dict, model: dict):
    """Returns (best_code, scores_dict) for a user."""
    user_ans = answers.get(email, {})
    if not user_ans or not model:
        return None, {}
    scores = compute_llr_scores(user_ans, model)
    best = max(scores, key=scores.get)
    return best, scores

# ── Sheet 1: Type Analysis ─────────────────────────────────────────────────────
def build_type_analysis_sheet(ws, users, answers, model):
    ws.title = "ניתוח טיפוסים"

    cols = [
        ("מייל",                       30),
        ("שם",                          16),
        ("קוד מאבחן (פנינה)",           16),
        ("הצעה סטטיסטית",               16),
        ("סטטיסטי = מאבחן?",            16),
        ("ציון LLR מוביל",              14),
        ("כל ציוני LLR",                50),
    ]
    for ci, (h, w) in enumerate(cols, 1):
        _hdr(ws, 1, ci, h, w)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    for ri, u in enumerate(sorted(users, key=lambda x: x["email"]), 2):
        email    = u["email"]
        pdn_code = u["pdn_code"]
        name     = f"{u.get('first_name','')} {u.get('last_name','')}".strip()
        sugg, scores = get_suggestion(email, answers, model)

        if sugg and pdn_code == sugg:
            match_val  = "כן"
            match_fill = MATCH_FILL
        elif sugg:
            match_val  = "לא"
            match_fill = MISMATCH_FILL
        else:
            match_val  = "N/A"
            match_fill = EMPTY_FILL

        row_fill = ALT_FILL if ri % 2 == 0 else EMPTY_FILL
        top_score = max(scores.values()) if scores else 0
        all_scores_str = "  |  ".join(f"{c}: {v:.2f}" for c, v in sorted(scores.items(), key=lambda x: -x[1]))

        _cell(ws, ri, 1, email,         row_fill, "left")
        _cell(ws, ri, 2, name,          row_fill, "left")
        _cell(ws, ri, 3, pdn_code,      row_fill)
        _cell(ws, ri, 4, sugg or "N/A", row_fill)
        _cell(ws, ri, 5, match_val,     match_fill)
        _cell(ws, ri, 6, round(top_score, 2), row_fill)
        _cell(ws, ri, 7, all_scores_str, row_fill, "left")

# ── Sheet 2: Per-Code Summary ──────────────────────────────────────────────────
def build_summary_sheet(ws, users, answers, model):
    ws.title = "סיכום לפי קוד"

    cols = [
        ("קוד PDN",                     10),
        ("סה\"כ מקרים",                  14),
        ("סטטיסטי = מאבחן",              16),
        ("% התאמה סטטיסטי",             18),
    ]
    for ci, (h, w) in enumerate(cols, 1):
        _hdr(ws, 1, ci, h, w)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    code_stats = {c: {"total": 0, "stat_match": 0} for c in PDN_CODES}

    for u in users:
        code = u["pdn_code"]
        if code not in PDN_CODES:
            continue
        sugg, _ = get_suggestion(u["email"], answers, model)
        code_stats[code]["total"] += 1
        if sugg == code:
            code_stats[code]["stat_match"] += 1

    # Totals row data
    grand_total = sum(s["total"] for s in code_stats.values())
    grand_match = sum(s["stat_match"] for s in code_stats.values())

    for ri, code in enumerate(PDN_CODES, 2):
        s     = code_stats[code]
        total = s["total"]
        sm    = s["stat_match"]

        if total > 0:
            pct_str = f"{sm}/{total} = {sm/total*100:.1f}%"
            pct_fill = MATCH_FILL if sm / total >= 0.6 else MISMATCH_FILL
        else:
            pct_str  = "אין נתונים"
            pct_fill = EMPTY_FILL

        row_fill = ALT_FILL if ri % 2 == 0 else EMPTY_FILL
        _cell(ws, ri, 1, code,    SECTION_FILL)
        _cell(ws, ri, 2, total,   row_fill)
        _cell(ws, ri, 3, sm,      row_fill)
        _cell(ws, ri, 4, pct_str, pct_fill)

    # Grand total row
    total_ri = len(PDN_CODES) + 2
    gp = f"{grand_match}/{grand_total} = {grand_match/grand_total*100:.1f}%" if grand_total else "-"
    _cell(ws, total_ri, 1, "סה\"כ",  PatternFill("solid", fgColor="4F4F4F"))
    ws.cell(total_ri, 1).font = Font(bold=True, color="FFFFFF")
    _cell(ws, total_ri, 2, grand_total, PatternFill("solid", fgColor="D9D9D9"))
    _cell(ws, total_ri, 3, grand_match, PatternFill("solid", fgColor="D9D9D9"))
    fill_total = MATCH_FILL if grand_total and grand_match/grand_total >= 0.6 else MISMATCH_FILL
    _cell(ws, total_ri, 4, gp, fill_total)

# ── main ───────────────────────────────────────────────────────────────────────
def main():
    if not DATA_JSON.exists():
        print(f"Production data not found at {DATA_JSON}")
        print("Run: TOKEN=... curl -X POST https://pdn-chat.onrender.com/pdn-admin/api/pdn-analysis/data ...")
        return

    print("Loading production data...")
    with open(DATA_JSON) as f:
        data = json.load(f)

    users   = data.get("users", [])
    answers = data.get("answers", {})

    labeled = [u for u in users if u.get("pdn_code", "") in PDN_CODES]
    print(f"Total labeled users: {len(labeled)}")
    from collections import Counter
    dist = Counter(u["pdn_code"] for u in labeled)
    for code in PDN_CODES:
        print(f"  {code}: {dist.get(code, 0)}")

    print("Training LLR model...")
    model = train_llr_model(labeled, answers)
    print(f"Model trained on {len(labeled)} users")

    if not EXCEL.exists():
        EXCEL.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        # remove default sheet
        wb.remove(wb.active)
    else:
        wb = openpyxl.load_workbook(EXCEL)

    # Remove existing statistics sheets
    for name in ["ניתוח טיפוסים", "סיכום לפי קוד"]:
        if name in wb.sheetnames:
            del wb[name]

    ws1 = wb.create_sheet("ניתוח טיפוסים")
    ws2 = wb.create_sheet("סיכום לפי קוד")

    print("Building type analysis sheet...")
    build_type_analysis_sheet(ws1, labeled, answers, model)

    print("Building summary sheet...")
    build_summary_sheet(ws2, labeled, answers, model)

    wb.save(EXCEL)
    print(f"\nSaved: {EXCEL}")

    # Quick accuracy check
    correct = sum(
        1 for u in labeled
        if get_suggestion(u["email"], answers, model)[0] == u["pdn_code"]
    )
    print(f"Overall LLR accuracy: {correct}/{len(labeled)} = {correct/len(labeled)*100:.1f}%")


if __name__ == "__main__":
    main()
