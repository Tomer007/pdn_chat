#!/usr/bin/env python3
"""
Rebuild Q38-Q56 and Q57-Q61 statistics sheets using full JSON answer data.
Matches JSON files to PDN codes via:
1. Email suffix e.g. tomergur+E1@gmail.com -> E1
2. Excel Q38-Q42 fingerprint matching
3. Excel direct match for known users
"""

import json, glob, os, re
from collections import defaultdict, Counter
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

PDN_CODES = ['E1', 'E5', 'E9', 'A3', 'A7', 'A11', 'T4', 'T8', 'T12', 'P2', 'P6', 'P10']

# ===========================
# LOAD EXCEL & BUILD FINGERPRINTS
# ===========================
wb = openpyxl.load_workbook('/Users/tomer.gur/dev-tools/pdn_chat/docs/combined_matrix_report.xlsx')
ws_answers = wb['תשובות משתמשים']
header = [cell.value for cell in ws_answers[1]]
q38_idx = header.index('Q38')

# fingerprint = tuple of Q38-Q56 answers
fp_to_pdn = defaultdict(list)
for row in ws_answers.iter_rows(min_row=2, values_only=True):
    user_str = row[0]
    if not user_str:
        continue
    parts = [p.strip() for p in str(user_str).split('|')]
    pdn_code = parts[0]
    fp = tuple(row[q38_idx:q38_idx + 19])  # Q38-Q56 all 19 answers
    fp_to_pdn[fp].append(pdn_code)

# ===========================
# LOAD JSON FILES & MATCH
# ===========================
json_files = glob.glob('/Users/tomer.gur/dev-tools/pdn_chat/saved_results/**/*_answers.json', recursive=True)

data_by_code = defaultdict(lambda: {
    'q38_q56': defaultdict(list),
    'q57_q61_pos': defaultdict(lambda: defaultdict(lambda: defaultdict(int))),
    'q57_q61_first': defaultdict(list),
    'count': 0
})

matched = 0
no_q38 = 0
no_match = 0

for jf in json_files:
    with open(jf) as f:
        try:
            data = json.load(f)
        except:
            continue

    # Skip if no Q38
    if '38' not in data:
        no_q38 += 1
        continue

    meta = data.get('metadata', {})
    email = meta.get('email', '').lower().strip()

    # Method 1: email encodes PDN code e.g. tomergur+E1@gmail.com
    pdn_code = None
    m = re.search(r'\+([EATP]\d+)@', email, re.IGNORECASE)
    if m:
        code = m.group(1).upper()
        if code in PDN_CODES:
            pdn_code = code

    # Method 2: fingerprint match from Excel
    if not pdn_code:
        fp = tuple(
            data[str(q)].get('selected_option_code') if isinstance(data.get(str(q)), dict) else None
            for q in range(38, 57)
        )
        candidates = fp_to_pdn.get(fp, [])
        if len(candidates) == 1:
            pdn_code = candidates[0]
        elif len(candidates) > 1:
            # Multiple - use most common
            pdn_code = Counter(candidates).most_common(1)[0][0]

    if not pdn_code:
        no_match += 1
        continue

    pdn_code = pdn_code.upper()
    if pdn_code not in PDN_CODES:
        continue

    matched += 1
    data_by_code[pdn_code]['count'] += 1

    # Q38-Q56 binary answers
    for q_num in range(38, 57):
        q_str = str(q_num)
        if q_str in data and isinstance(data[q_str], dict):
            ans = data[q_str].get('selected_option_code')
            if ans:
                data_by_code[pdn_code]['q38_q56'][q_num].append(ans)

    # Q57-Q61 full rankings
    for q_num in range(57, 62):
        q_str = str(q_num)
        if q_str not in data:
            continue
        q_data = data[q_str]
        if not isinstance(q_data, dict):
            continue
        ranking = q_data.get('ranking', {})
        if not ranking or not isinstance(ranking, dict):
            continue
        # pos -> trait counts
        for trait, pos in ranking.items():
            if trait in ['T', 'A', 'E', 'P'] and isinstance(pos, int) and 1 <= pos <= 4:
                data_by_code[pdn_code]['q57_q61_pos'][q_num][pos][trait] += 1
        # First choice
        first = min(
            ((t, p) for t, p in ranking.items() if t in ['T', 'A', 'E', 'P'] and isinstance(p, int)),
            key=lambda x: x[1],
            default=(None, 99)
        )
        if first[0]:
            data_by_code[pdn_code]['q57_q61_first'][q_num].append(first[0])

print(f'Matched: {matched} | No Q38: {no_q38} | No match: {no_match}')
for code in PDN_CODES:
    n = data_by_code[code]['count']
    if n > 0:
        print(f'  {code}: N={n}')

# ===========================
# STYLES
# ===========================
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
bold_font = Font(bold=True)
thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def pct(cnt, total):
    if total == 0: return ''
    return f'{round(cnt * 100 / total)}%'

def make_header(ws, headers):
    for col, val in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=val)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin
    ws.row_dimensions[1].height = 45

# ===========================
# Q38-Q56 SHEET
# ===========================
Q38_56_LABELS = {
    38: "בדיקה לעומק לפני נטילת סיכון",
    39: "קושי עם מצבי חוסר ודאות",
    40: "קבלת החלטות במהירות",
    41: "קושי לקבל הנחיות מאחרים",
    42: "נוחות להוביל אחרים",
    43: "מוחצנות / מופנמות - ילדות",
    44: "מוחצנות / מופנמות - בגרות",
    45: "הובלה / הצטרפות - ילדות",
    46: "הובלה / הצטרפות - בגרות",
    47: "דברנות / שתקנות - ילדות",
    48: "דברנות / שתקנות - בגרות",
    49: "עימותים / ריצוי - ילדות",
    50: "עימותים / ריצוי - בגרות",
    51: "עמידה בזמנים - ילדות",
    52: "עמידה בזמנים - בגרות",
    53: "סדר / בלגן - ילדות",
    54: "סדר / בלגן - בגרות",
    55: "מאופקות / נועזות - ילדות",
    56: "מאופקות / נועזות - בגרות",
}

if 'Q38-Q56 סטטיסטיקה' in wb.sheetnames:
    del wb['Q38-Q56 סטטיסטיקה']
ws38 = wb.create_sheet('Q38-Q56 סטטיסטיקה')

make_header(ws38, ['שאלה'] + [f'{c}\n(N={data_by_code[c]["count"]})' for c in PDN_CODES])

for ri, q_num in enumerate(range(38, 57), 2):
    label = Q38_56_LABELS[q_num]
    c = ws38.cell(row=ri, column=1, value=f'Q{q_num} - {label}')
    c.border = thin
    c.alignment = Alignment(wrap_text=True, vertical='center')

    for ci, code in enumerate(PDN_CODES, 2):
        answers = data_by_code[code]['q38_q56'].get(q_num, [])
        total = len(answers)
        if total == 0:
            cell = ws38.cell(row=ri, column=ci, value='-')
        else:
            counts = Counter(answers)
            lines = []
            for trait, cnt in sorted(counts.items(), key=lambda x: -x[1]):
                lines.append(f'{trait}: {pct(cnt, total)} ({cnt}/{total})')
            cell = ws38.cell(row=ri, column=ci, value='\n'.join(lines))
        cell.border = thin
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    ws38.row_dimensions[ri].height = 35

ws38.column_dimensions['A'].width = 32
for col in range(2, 14):
    ws38.column_dimensions[get_column_letter(col)].width = 14

print('Q38-Q56 sheet rebuilt')

# ===========================
# Q57-Q61 SHEET
# ===========================
Q57_61_LABELS = {
    57: "דירוג: אסרטיבי / אכפתי / שיטתי / שופע רעיונות",
    58: "דירוג: מוביל / תומך / זהיר / רעיוניסט",
    59: "דירוג: ממוקד מטרה / נותן / שיטתי / אופטימי",
    60: "מה היית רוצה לקבל מהאבחון",
    61: "מה חשוב לך שהאבחון יספק",
}

if 'Q57-Q61 סטטיסטיקה' in wb.sheetnames:
    del wb['Q57-Q61 סטטיסטיקה']
ws57 = wb.create_sheet('Q57-Q61 סטטיסטיקה')

make_header(ws57, ['שאלה'] + [f'{c}\n(N={data_by_code[c]["count"]})' for c in PDN_CODES])

for ri, q_num in enumerate(range(57, 62), 2):
    label = Q57_61_LABELS[q_num]
    c = ws57.cell(row=ri, column=1, value=f'Q{q_num}\n{label}')
    c.border = thin
    c.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    c.alignment = Alignment(wrap_text=True, vertical='center')

    for ci, code in enumerate(PDN_CODES, 2):
        first_choices = data_by_code[code]['q57_q61_first'].get(q_num, [])
        pos_data = data_by_code[code]['q57_q61_pos'].get(q_num, {})
        total = len(first_choices)

        if total == 0:
            cell = ws57.cell(row=ri, column=ci, value='-')
        else:
            lines = []
            # First choice % per trait (sorted by most common)
            fc_counts = Counter(first_choices)
            for trait in sorted(fc_counts, key=lambda t: -fc_counts[t]):
                lines.append(f'{trait}: {pct(fc_counts[trait], total)} ({fc_counts[trait]})')
            
            # Position distribution
            lines.append('')
            for pos in [1, 2, 3, 4]:
                pd = pos_data.get(pos, {})
                if pd:
                    pos_str = f'מקום {pos}: ' + '  '.join(f'{t}={c}' for t, c in sorted(pd.items()))
                    lines.append(pos_str)
            
            cell = ws57.cell(row=ri, column=ci, value='\n'.join(lines).strip())
        cell.border = thin
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    ws57.row_dimensions[ri].height = 90

ws57.column_dimensions['A'].width = 38
for col in range(2, 14):
    ws57.column_dimensions[get_column_letter(col)].width = 16

print('Q57-Q61 sheet rebuilt')

# ===========================
# SAVE & VALIDATE
# ===========================
wb.save('/Users/tomer.gur/dev-tools/pdn_chat/docs/combined_matrix_report.xlsx')
print('\nSaved. Final sheets:', wb.sheetnames)

print('\n=== VALIDATION ===')
wb2 = openpyxl.load_workbook('/Users/tomer.gur/dev-tools/pdn_chat/docs/combined_matrix_report.xlsx')

# REQ 1
ws_org = wb2['גיליון תשובות מאורגנים']
headers = [cell.value for cell in ws_org[1]]
print(f'REQ 1 - Organized sheet: headers={headers}, rows={ws_org.max_row-1}')
print(f'  PASS: {headers == ["Name", "UID", "Verified PDN Code"] and ws_org.max_row > 1}')

# REQ 2
ws_q38 = wb2['Q38-Q56 סטטיסטיקה']
rows_38 = ws_q38.max_row - 1
cols_38 = ws_q38.max_column
print(f'REQ 2 - Q38-Q56: rows={rows_38} (need 19), cols={cols_38} (need 13)')
print(f'  Sample E5/Q38: {ws_q38.cell(row=2, column=3).value}')
print(f'  PASS: {rows_38 == 19 and cols_38 == 13}')

# REQ 3
ws_q57 = wb2['Q57-Q61 סטטיסטיקה']
rows_57 = ws_q57.max_row - 1
cols_57 = ws_q57.max_column
print(f'REQ 3 - Q57-Q61: rows={rows_57} (need 5), cols={cols_57} (need 13)')
for code in PDN_CODES:
    col_idx = PDN_CODES.index(code) + 2
    val = ws_q57.cell(row=2, column=col_idx).value
    if val and val != '-':
        print(f'  {code}/Q57: {str(val)[:80]}')
print(f'  PASS: {rows_57 == 5 and cols_57 == 13}')
