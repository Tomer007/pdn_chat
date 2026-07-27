import csv
import hmac
import json
import logging
import os
import re
import secrets
import threading
import time
from datetime import datetime, timedelta
from functools import wraps

from flask import Blueprint, request, render_template, jsonify, current_app, send_file, abort, make_response
from pathlib import Path
from werkzeug.exceptions import HTTPException

from ..utils.answer_storage import load_answers
from ..utils.csv_metadata_handler import UserMetadataHandler
from ..utils.email_sender import send_pdn_code_email, send_binat_invite_email, send_coupon_invite_email
from ..utils.pdn_calculator import calculate_pdn_code, check_verification_needed
from ..utils.pdn_file_path import PDNFilePath
from ..utils.conversation_stats import conversation_stats
from ..version import VERSION, RELEASE_DATE, RELEASE_NOTES
from ..pdn_chat_ai.user_manager import get_user_manager
from .coupon_manager import get_coupon_manager


# Configure logging
logger = logging.getLogger(__name__)

# Add expiration tracking
SESSION_TIMEOUT = timedelta(hours=2)



# Create blueprint
pdn_admin_bp = Blueprint('pdn_admin', __name__,
                         template_folder='templates',
                         static_folder='../static')

# Admin sessions storage (in production, use Redis or database)
admin_sessions = {}  # session_token -> user_info
MAX_ADMIN_SESSIONS = 50  # Prevent unbounded growth

def create_session(email):
    # Remove any existing sessions for this email
    for token, session in list(admin_sessions.items()):
        if session.get("email") == email:
            del admin_sessions[token]
            logger.info("Removed old session for %s", email)

    # Enforce max sessions limit
    cleanup_expired_sessions()
    if len(admin_sessions) >= MAX_ADMIN_SESSIONS:
        oldest = min(admin_sessions, key=lambda k: admin_sessions[k].get('login_time', datetime.min))
        del admin_sessions[oldest]

    token = secrets.token_urlsafe(32)
    now = datetime.now()
    admin_sessions[token] = {
        "email": email,
        "username": email,
        "login_time": now,
        "expires_at": now + SESSION_TIMEOUT
    }
    logger.info("Created new session for %s: %s", email, token)
    return token


def verify_session(session_token: str):
    cleanup_expired_sessions()

    if not session_token or session_token not in admin_sessions:
        abort(make_response(jsonify({"error": "Invalid or expired session"}), 401))

    session = admin_sessions[session_token]
    if datetime.now() > session["expires_at"]:
        del admin_sessions[session_token]
        abort(make_response(jsonify({"error": "Session expired"}), 401))

    # Sliding window: refresh expiry on every successful verification
    session["expires_at"] = datetime.now() + SESSION_TIMEOUT

    return session

# Clean up expired sessions periodically
def cleanup_expired_sessions():
    now = datetime.now()
    expired = [token for token, session in admin_sessions.items()
               if now > session["expires_at"]]
    for token in expired:
        del admin_sessions[token]


def require_admin_session(f):
    """Decorator that verifies admin session token from query params."""
    @wraps(f)
    def decorated(*args, **kwargs):
        verify_session(request.args.get('session_token'))
        return f(*args, **kwargs)
    return decorated


# Regex for validating coupon codes in URL paths (1-20 alphanumeric chars)
_COUPON_CODE_PATTERN = re.compile(r'^[A-Za-z0-9]{1,20}$')


def _validate_coupon_code_param(code: str) -> bool:
    """Validate that a coupon code URL parameter is safe and well-formed."""
    return bool(_COUPON_CODE_PATTERN.match(code))

_metadata_cache = {'data': None, 'timestamp': 0}
_METADATA_CACHE_TTL = 60  # seconds

# Error tracker for health monitoring
_error_tracker = {'count_24h': 0, 'last_error': None, 'last_error_time': None, 'errors': []}


def _track_error(message: str):
    """Track an error for health monitoring."""
    from datetime import datetime
    now = datetime.now()
    _error_tracker['count_24h'] += 1
    _error_tracker['last_error'] = message[:100]
    _error_tracker['last_error_time'] = now.strftime('%H:%M:%S')
    _error_tracker['errors'].append({'msg': message[:80], 'time': now.strftime('%H:%M')})
    # Keep only last 20 errors
    if len(_error_tracker['errors']) > 20:
        _error_tracker['errors'] = _error_tracker['errors'][-20:]


class _ErrorTrackingHandler(logging.Handler):
    """Custom logging handler that counts errors for health monitoring."""
    def emit(self, record):
        if record.levelno >= logging.ERROR:
            _track_error(record.getMessage())


# Attach the error tracking handler to the admin logger
logger.addHandler(_ErrorTrackingHandler())

def load_user_metadata():
    """
    Load user metadata from the CSV file and JSON files.
    
    Returns:
        List of dictionaries containing user metadata
    """
    try:
        now = time.time()
        if _metadata_cache['data'] is not None and (now - _metadata_cache['timestamp']) < _METADATA_CACHE_TTL:
            return _metadata_cache['data']

        csv_file_path = Path(os.getenv('SAVED_RESULTS_DIR', 'saved_results')) / 'user_metadata.csv'
        logger.info("CSV file path: %s", csv_file_path)
        if not csv_file_path.exists():
            logger.warning("user_metadata.csv file not found")
            return []

        metadata_list = []
        csv_metadata_handler = UserMetadataHandler()

        with open(csv_file_path, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                # Skip empty rows
                if not row.get("Email", "").strip():
                    continue

                email = row.get("Email", "").strip()

                # Load additional metadata from JSON file
                json_metadata = {}
                try:
                    questionnaire_data = csv_metadata_handler.get_user_files(email, "answers")
                    if questionnaire_data and 'metadata' in questionnaire_data:
                        json_metadata = questionnaire_data['metadata']
                except Exception as e:
                    logger.warning("Could not load JSON metadata for %s: %s", email, e)

                # Convert CSV column names to the expected format and merge with JSON metadata
                user_data = {
                    "user_id": (row.get("User ID") or "").strip(),
                    "email": email,
                    "date": (row.get("Date") or "").strip(),
                    "pdn_code": (row.get("PDN Code") or "").strip(),
                    "pdn_voice_code": (row.get("PDN Voice Code") or "").strip(),
                    "diagnose_pdn_code": (row.get("Diagnose PDN Code") or "").strip(),
                    "diagnose_comments": (row.get("Diagnose Comments") or "").strip(),
                    "pdn_update_comments": (row.get("PDN Update Comments") or "").strip(),
                    # Load from JSON metadata if available, otherwise use CSV or defaults
                    "first_name": (json_metadata.get("first_name") or row.get("First Name") or "").strip(),
                    "last_name": (json_metadata.get("last_name") or row.get("Last Name") or "").strip(),
                    "phone": (json_metadata.get("phone") or row.get("Phone") or "").strip(),
                    "native_language": (
                                json_metadata.get("native_language") or json_metadata.get("mother_language") or row.get(
                            "Native Language") or "").strip(),
                    "gender": (json_metadata.get("gender") or row.get("Gender") or "").strip(),
                    "education_level": (
                                json_metadata.get("education_level") or json_metadata.get("education") or row.get(
                            "Education Level") or "").strip(),
                    "job_title": (json_metadata.get("job_title") or row.get("Job Title") or "").strip(),
                    "birth_year": (json_metadata.get("birth_year") or row.get("Birth Year") or "").strip(),
                    "coupon_code": (json_metadata.get("coupon_code") or "").strip(),
                    "link_to_user": f"/user/{email}",
                    "questionnaire": f"/api/user/questionnaire/{email}",
                    "voice": f"/api/user/voice/{email}"
                }

                # Calculate needs_verification from user's answers
                needs_verification = False
                stage_e_override = False
                dominant_before_stage_e = None
                confidence_score = None
                try:
                    user_answers = load_answers(email)
                    if user_answers:
                        calc_result = calculate_pdn_code(user_answers, return_details=True, user_id=email)
                        if isinstance(calc_result, dict):
                            needs_verification = calc_result.get('needs_verification', False)
                            stage_e_override = calc_result.get('stage_e_override', False)
                            dominant_before_stage_e = calc_result.get('dominant_before_stage_e')
                            # Extract confidence from calculation_details Final stage
                            details = calc_result.get('calculation_details', [])
                            for d in details:
                                if d.get('stage') == 'Final':
                                    confidence_score = d.get('confidence_score')
                                    break
                        else:
                            needs_verification = False
                except Exception as e:
                    logger.debug("Could not calculate verification for %s: %s", email, e)
                user_data["needs_verification"] = needs_verification
                user_data["stage_e_override"] = stage_e_override
                user_data["dominant_before_stage_e"] = dominant_before_stage_e
                if confidence_score is not None:
                    user_data["confidence_score"] = confidence_score

                metadata_list.append(user_data)

        logger.info("Loaded %d user records from CSV and JSON", len(metadata_list))
        _metadata_cache['data'] = metadata_list
        _metadata_cache['timestamp'] = time.time()
        return metadata_list

    except Exception as e:
        logger.error("Error loading user metadata from CSV: %s", e)
        return []

@pdn_admin_bp.route('/')
def admin_login_page():
    """Admin login page"""
    return render_template("admin_login.html")


@pdn_admin_bp.route('/dashboard')
def admin_dashboard_page():
    """Admin dashboard page"""
    return render_template("admin_dashboard.html")


@pdn_admin_bp.route('/login', methods=['POST'])
def admin_login():
    """Admin login endpoint"""

    try:
        login_data = request.get_json()
        email = login_data.get('email', '')
        password = login_data.get('password', '')

        expected = current_app.config.get('ADMIN_PASSWORD', 'jclazvbdn')
        logger.info("Admin login attempt: email=%s, password_len=%d, expected_len=%d", email, len(password), len(expected))

        if hmac.compare_digest(password, expected):
            session_token = create_session(email)
            return jsonify({
                "success": True,
                "message": "Login successful",
                "session_token": session_token
            })

        # Log failed login attempt
        logger.warning("Failed admin login attempt for email: %s", email)

        return jsonify({"error": "Invalid credentials"}), 401

    except Exception as e:
        logger.error("Login error: %s", e)
        return jsonify({"error": "Login failed"}), 400


@pdn_admin_bp.route('/logout')
def admin_logout():
    """Admin logout endpoint"""

    admin_sessions.pop(request.args.get('session_token'), None)
    cleanup_expired_sessions()
    return jsonify({"success": True, "message": "Logout successful"})

def _format_user(s, user_type):
    """Helper to format user session data"""
    fmt = "%d/%m/%Y %H:%M:%S"
    user = {
        "email": s.get("email", s.get("user_id", "unknown")),
        "login_time": s["login_time"].strftime(fmt) if "login_time" in s else "N/A",
        "type": user_type
    }
    if "expires_at" in s:
        user["expires_at"] = s["expires_at"].strftime(fmt)
    return user

@pdn_admin_bp.route('/logged-in-users')
def get_logged_in_users():
    """Get list of all logged-in users from all apps"""
    try:
        verify_session(request.args.get('session_token'))
        
        users = [_format_user(s, "admin") for s in admin_sessions.values()]
        
        # Diagnosis users
        try:
            from ..pdn_diagnose.diagnosis_routes import active_sessions
            users.extend(_format_user(s, "diagnosis") for s in active_sessions.values())
            logger.debug("Loaded %d diagnosis sessions", len(active_sessions))
        except (ImportError, AttributeError) as e:
            logger.debug("Could not load diagnosis sessions: %s", e)
        
        # Chat AI users
        try:
            from ..pdn_chat_ai.chat_routes import chat_sessions
            users.extend(_format_user(s, "chat_ai") for s in chat_sessions.values())
            logger.debug("Loaded %d chat sessions", len(chat_sessions))
        except (ImportError, AttributeError) as e:
            logger.debug("Could not load chat sessions: %s", e)
        
        return jsonify({"users": users, "count": len(users)})
    except Exception:
        return jsonify({"error": "Unauthorized"}), 401


@pdn_admin_bp.route('/metadata/csv')
def get_metadata_csv():

    try:
        verify_session(request.args.get('session_token'))
    except Exception as e:
        logger.error("Session verification failed: %s", e)
        return jsonify({"error": "Unauthorized"}), 401

    return jsonify({"data": load_user_metadata()})


@pdn_admin_bp.route('/send_algorithm_report', methods=['POST'])
def send_algorithm_report():
    """Send PDN algorithm report HTML as email attachment to admin."""
    try:
        verify_session(request.args.get('session_token'))
    except Exception as e:
        logger.error("Session verification failed: %s", e)
        return jsonify({"error": "Unauthorized"}), 401

    try:
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.application import MIMEApplication
        from ..utils.email_sender import EmailConfig, send_email_via_smtp

        recipient = 'tomergur@gmail.com'
        report_path = Path(current_app.root_path).parent / 'docs' / 'pdn_algorithm_report.html'

        if not report_path.exists():
            return jsonify({"error": "Report file not found"}), 404

        msg = MIMEMultipart()
        msg['From'] = EmailConfig.FROM_EMAIL
        msg['To'] = recipient
        msg['Subject'] = 'פירוט חישוב קוד PDN — דוח אלגוריתם'

        body = MIMEText('מצורף דוח מפורט על אלגוריתם חישוב קוד PDN.\n\nנשלח ממערכת הניהול של PDN.', 'plain', 'utf-8')
        msg.attach(body)

        with open(report_path, 'rb') as f:
            attachment = MIMEApplication(f.read(), _subtype='html')
            attachment.add_header('Content-Disposition', 'attachment', filename='pdn_algorithm_report.html')
            msg.attach(attachment)

        success = send_email_via_smtp(msg)
        if success:
            logger.info("Algorithm report sent to %s", recipient)
            return jsonify({"success": True, "message": f"דוח נשלח בהצלחה ל-{recipient}"})
        else:
            return jsonify({"error": "Failed to send email"}), 500

    except Exception as e:
        logger.error("Error sending algorithm report: %s", e)
        return jsonify({"error": str(e)}), 500


@pdn_admin_bp.route('/send_calculation_report', methods=['POST'])
def send_calculation_report():
    """Send PDN calculation details (from recalculation modal) as HTML email to admin."""
    try:
        verify_session(request.args.get('session_token'))
    except Exception as e:
        logger.error("Session verification failed: %s", e)
        return jsonify({"error": "Unauthorized"}), 401

    try:
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from ..utils.email_sender import EmailConfig, send_email_via_smtp

        data = request.get_json()
        user_email = data.get('email', 'unknown')
        html_content = data.get('html_content', '')

        if not html_content:
            return jsonify({"error": "No content to send"}), 400

        recipient = 'tomergur@gmail.com'

        msg = MIMEMultipart('alternative')
        msg['From'] = EmailConfig.FROM_EMAIL
        msg['To'] = recipient
        msg['Subject'] = f'פירוט חישוב קוד PDN — {user_email}'

        # Wrap the content in a full HTML document for email
        full_html = f"""<!DOCTYPE html>
<html lang="he" dir="rtl">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"></head>
<body style="font-family: Inter, -apple-system, sans-serif; background: #f8fafc; padding: 24px; direction: rtl;">
<div style="max-width: 700px; margin: 0 auto; background: white; border-radius: 16px; padding: 32px; box-shadow: 0 4px 24px rgba(0,0,0,0.06);">
<h2 style="color: #0b2e6b; font-size: 18px; margin-bottom: 24px;">פירוט חישוב קוד PDN — {user_email}</h2>
{html_content}
</div>
</body></html>"""

        msg.attach(MIMEText(full_html, 'html', 'utf-8'))

        success = send_email_via_smtp(msg)
        if success:
            logger.info("Calculation report for %s sent to %s", user_email, recipient)
            return jsonify({"success": True, "message": "דוח חישוב נשלח בהצלחה"})
        else:
            return jsonify({"error": "Failed to send email"}), 500

    except Exception as e:
        logger.error("Error sending calculation report: %s", e)
        return jsonify({"error": str(e)}), 500


@pdn_admin_bp.route('/compress_old_audio', methods=['POST'])
def compress_old_audio():
    """Compress WAV audio files older than 30 days to MP3."""
    try:
        verify_session(request.args.get('session_token'))
    except Exception as e:
        logger.error("Session verification failed: %s", e)
        return jsonify({"error": "Unauthorized"}), 401

    try:
        import subprocess
        saved_results_dir = Path(os.getenv('SAVED_RESULTS_DIR', 'saved_results'))

        if not saved_results_dir.exists():
            return jsonify({"error": "saved_results directory not found"}), 404

        # Calculate current storage usage
        total_size = sum(f.stat().st_size for f in saved_results_dir.rglob('*') if f.is_file())
        wav_total = sum(f.stat().st_size for f in saved_results_dir.rglob('*.wav'))
        mp3_total = sum(f.stat().st_size for f in saved_results_dir.rglob('*.mp3'))

        # Find WAV files older than 30 days
        cutoff = time.time() - (30 * 86400)
        wav_files = []
        for wav_file in saved_results_dir.rglob('*.wav'):
            if wav_file.stat().st_mtime < cutoff:
                wav_files.append(wav_file)

        storage_info = {
            "total_mb": round(total_size / (1024 * 1024), 1),
            "wav_mb": round(wav_total / (1024 * 1024), 1),
            "mp3_mb": round(mp3_total / (1024 * 1024), 1),
            "wav_old_count": len(wav_files),
            "wav_old_mb": round(sum(f.stat().st_size for f in wav_files) / (1024 * 1024), 1),
            "disk_limit_mb": 1024
        }

        # If GET-like check (no actual compression requested)
        data = request.get_json(silent=True) or {}
        if data.get('check_only'):
            return jsonify({"success": True, "storage": storage_info})

        if not wav_files:
            return jsonify({"success": True, "message": "אין קבצי WAV ישנים מ-30 יום", "compressed": 0, "saved_mb": 0, "storage": storage_info})

        success = 0
        failed = 0
        bytes_saved = 0

        for wav_file in wav_files:
            mp3_file = wav_file.with_suffix('.mp3')
            wav_size = wav_file.stat().st_size

            try:
                result = subprocess.run(
                    ['ffmpeg', '-i', str(wav_file), '-b:a', '64k', '-y', str(mp3_file)],
                    capture_output=True, text=True, timeout=30
                )
                if result.returncode == 0 and mp3_file.exists():
                    mp3_size = mp3_file.stat().st_size
                    bytes_saved += wav_size - mp3_size
                    wav_file.unlink()  # Delete original WAV
                    success += 1
                else:
                    failed += 1
                    logger.warning("ffmpeg failed for %s: %s", wav_file.name, result.stderr[:100])
            except Exception as e:
                failed += 1
                logger.warning("Error compressing %s: %s", wav_file.name, e)

        saved_mb = round(bytes_saved / (1024 * 1024), 1)
        logger.info("Audio compression complete: %d compressed, %d failed, %.1f MB saved", success, failed, saved_mb)

        # Recalculate storage after compression
        total_after = sum(f.stat().st_size for f in saved_results_dir.rglob('*') if f.is_file())
        storage_info["total_mb"] = round(total_after / (1024 * 1024), 1)

        return jsonify({
            "success": True,
            "message": f"דחיסה הושלמה: {success} קבצים דוחסו, חסכון {saved_mb} MB",
            "compressed": success,
            "failed": failed,
            "saved_mb": saved_mb,
            "storage": storage_info
        })

    except Exception as e:
        logger.error("Error in audio compression: %s", e)
        return jsonify({"error": str(e)}), 500



def remove_none_keys(obj):
    """Recursively remove None keys from dicts/lists."""
    if isinstance(obj, dict):
        return {k: remove_none_keys(v) for k, v in obj.items() if k is not None}
    elif isinstance(obj, list):
        return [remove_none_keys(item) for item in obj]
    else:
        return obj


@pdn_admin_bp.route('/health_status')
def get_health_status():
    """Get system health status for admin dashboard."""
    try:
        verify_session(request.args.get('session_token'))
    except Exception as e:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        import psutil
        saved_results_dir = Path(os.getenv('SAVED_RESULTS_DIR', 'saved_results'))

        # CPU and Memory from psutil
        cpu_percent = psutil.cpu_percent(interval=0.5)
        memory = psutil.virtual_memory()

        # Storage
        storage_total = 0
        if saved_results_dir.exists():
            storage_total = sum(f.stat().st_size for f in saved_results_dir.rglob('*') if f.is_file())

        # Active sessions count
        active_count = len(admin_sessions)
        try:
            from ..pdn_diagnose.diagnosis_routes import active_sessions as diag_sessions
            active_count += len(diag_sessions)
        except Exception:
            pass

        # Error logs — count from internal error tracker
        error_count_24h = _error_tracker.get('count_24h', 0)
        last_error = _error_tracker.get('last_error', None)
        last_error_time = _error_tracker.get('last_error_time', None)

        # Uptime
        import datetime as dt
        boot_time = dt.datetime.fromtimestamp(psutil.boot_time())
        uptime_seconds = (dt.datetime.now() - boot_time).total_seconds()
        uptime_hours = int(uptime_seconds / 3600)

        health = {
            "status": "operational",
            "cpu_percent": round(cpu_percent, 1),
            "memory_used_mb": round(memory.used / (1024 * 1024)),
            "memory_total_mb": round(memory.total / (1024 * 1024)),
            "memory_percent": round(memory.percent, 1),
            "storage_used_mb": round(storage_total / (1024 * 1024), 1),
            "storage_limit_mb": 1024,
            "active_sessions": active_count,
            "uptime_hours": uptime_hours,
            "errors_24h": error_count_24h,
            "last_error": last_error,
            "last_error_time": last_error_time,
            "service_url": "https://pdn-chat.onrender.com",
            "region": "Frankfurt",
            "plan": "Starter"
        }

        # Determine overall status
        if cpu_percent > 90 or memory.percent > 95 or error_count_24h > 50:
            health["status"] = "critical"
        elif cpu_percent > 70 or memory.percent > 80 or error_count_24h > 10:
            health["status"] = "warning"

        return jsonify(health)

    except Exception as e:
        logger.error("Error getting health status: %s", e)
        return jsonify({"status": "unknown", "error": str(e)})


@pdn_admin_bp.route('/user/questionnaire/<email>')
def get_user_questionnaire(email):
    """Get user questionnaire data"""

    try:
        verify_session(request.args.get('session_token'))
    except Exception as e:
        logger.error("Session verification failed: %s", e)
        return jsonify({"error": "Unauthorized"}), 401

    try:
        csv_metadata_handler = UserMetadataHandler()
        questionnaire_data = csv_metadata_handler.get_user_files(email, "answers")

        if not questionnaire_data:
            logger.warning("No questionnaire data found for user: %s", email)
            return jsonify({"error": "User questionnaire not found"}), 404

        user_metadata = csv_metadata_handler.get_user_by_email(email)

        if user_metadata:
            questionnaire_data.setdefault('metadata', {}).update(user_metadata)
            logger.info("Loaded questionnaire for %s with User ID: %s", email, user_metadata.get('User ID', 'N/A'))
        else:
            questionnaire_data.setdefault('metadata', {'email': email, 'User ID': 'N/A'})
            logger.warning("No CSV metadata found for user: %s", email)

        return jsonify(remove_none_keys(questionnaire_data))

    except Exception as e:
        logger.error("Error loading questionnaire for %s: %s", email, e, exc_info=True)
        return jsonify({"error": f"Failed to load questionnaire: {str(e)}"}), 500


@pdn_admin_bp.route('/user/voice/<email>')
def get_user_voice(email):
    """Get user voice recording URL"""
    try:
        verify_session(request.args.get('session_token'))
    except Exception as e:
        logger.error("Session verification failed: %s", e)
        return jsonify({"error": "Unauthorized"}), 401

    try:
        pdn_file_path = PDNFilePath()
        voice_recordings = {}

        for question_num in ['question1', 'question2']:
            # Try WAV first, then MP3
            filename = pdn_file_path.find_user_file(email, f"{question_num}.wav")
            if not filename or not filename.exists():
                filename = pdn_file_path.find_user_file(email, f"{question_num}.mp3")
            if filename and filename.exists():
                try:
                    if filename.is_file() and filename.stat().st_size > 0:
                        voice_recordings[question_num] = {
                            'filename': str(filename),
                            'path': str(filename),
                            'exists': True
                        }
                except (OSError, IOError) as e:
                    logger.warning("Error accessing %s file %s: %s", question_num, filename, e)

        if not voice_recordings:
            return jsonify({"error": "User voice recording not found"}), 404

        return jsonify({
            "email": email,
            "voice_recordings": voice_recordings,
            "has_recordings": True
        })
    except Exception as e:
        logger.error("Error finding user metadata: %s", e)
        return jsonify({"error": "User not found"}), 404


@pdn_admin_bp.route('/user/diagnose/<email>', methods=['PUT'])
def update_user_diagnose(email):
    """Update user diagnose information"""

    try:
        verify_session(request.args.get('session_token'))
    except Exception as e:
        logger.error("Session verification failed: %s", e)

    try:
        diagnose_data = request.get_json()
        user_data = next((user for user in load_user_metadata() if user["email"] == email), None)

        if not user_data:
            return jsonify({"error": "User not found"}), 404

        diagnose_pdn_code = diagnose_data.get("diagnose_pdn_code", user_data.get("pdn_code", ""))
        diagnose_comments = diagnose_data.get("diagnose_comments", "")

        user_data["diagnose_pdn_code"] = diagnose_pdn_code
        user_data["diagnose_comments"] = diagnose_comments

        try:
            UserMetadataHandler().update_diagnose_code(email, diagnose_pdn_code, diagnose_comments)
            logger.info("Successfully updated CSV with diagnose info for %s", email)
        except Exception as csv_error:
            logger.warning("Failed to update CSV with diagnose info: %s", csv_error)

        return jsonify({
            "success": True,
            "message": "Diagnose updated successfully",
            "user": user_data
        })
    except Exception as e:
        logger.error("Error updating diagnose: %s", e)
        return jsonify({"error": "Failed to update diagnose"}), 400

@pdn_admin_bp.route('/user/send_email/<email>', methods=['POST'])
def send_user_email(email):
    """Send PDN report email to user"""

    try:
        user_answers = load_answers(email)
        if not user_answers:
            return jsonify({"error": "User answers not found"}), 404

        # Ensure metadata.email is set (some older records may be missing it)
        if 'metadata' not in user_answers:
            user_answers['metadata'] = {}
        if not user_answers['metadata'].get('email'):
            user_answers['metadata']['email'] = email

        calculation_result = calculate_pdn_code(user_answers, user_id=email)

        if isinstance(calculation_result, dict):
            pdn_code = calculation_result['pdn_code']
            needs_verification = calculation_result.get('needs_verification', False)
        else:
            pdn_code = calculation_result
            needs_verification = False

        if not pdn_code:
            return jsonify({"error": "Could not calculate PDN code"}), 400

        logger.info("send_email PDN code: %s for user %s, needs_verification: %s", pdn_code, email, needs_verification)

        if not send_pdn_code_email(user_answers, pdn_code):
            return jsonify({"error": "Failed to send email"}), 500

        return jsonify({
            "success": True,
            "message": f"Email sent successfully to {email}",
            "pdn_code": pdn_code,
            "needs_verification": needs_verification
        })

    except Exception as e:
        logger.error("Error sending email: %s", e)
        return jsonify({"error": f"Error sending email: {str(e)}"}), 500


@pdn_admin_bp.route('/user/send_binat_invite/<email>', methods=['POST'])
def send_binat_invite(email):
    """Send Binat chat invitation email to user"""

    try:
        # Get user's first name from metadata
        csv_handler = UserMetadataHandler()
        questionnaire_data = csv_handler.get_user_files(email, "answers")

        first_name = ''
        if questionnaire_data and 'metadata' in questionnaire_data:
            first_name = questionnaire_data['metadata'].get('first_name', '')

        if not first_name:
            user_data = csv_handler.get_user_by_email(email)
            if user_data:
                first_name = user_data.get('First Name', '')

        if not send_binat_invite_email(email, first_name):
            return jsonify({"error": "Failed to send Binat invite email"}), 500

        return jsonify({
            "success": True,
            "message": f"Binat invite sent successfully to {email}"
        })

    except Exception as e:
        logger.error("Error sending Binat invite: %s", e)
        return jsonify({"error": f"Error sending Binat invite: {str(e)}"}), 500


@pdn_admin_bp.route('/user/recalculate_pdn/<email>', methods=['POST'])
def recalculate_user_pdn(email):
    """Recalculate PDN code for a user"""

    try:
        user_answers = load_answers(email)
        if not user_answers:
            return jsonify({"error": "User answers not found"}), 404

        calculation_result = calculate_pdn_code(user_answers, return_details=True, user_id=email)

        if isinstance(calculation_result, dict):
            pdn_code = calculation_result['pdn_code']
            calculation_details = calculation_result.get('calculation_details')
            needs_verification = calculation_result.get('needs_verification', False)
            confidence_score = calculation_result.get('confidence_score', 0)
        else:
            pdn_code = calculation_result
            calculation_details = None
            needs_verification = False
            confidence_score = 0

        if not pdn_code:
            return jsonify({"error": "Could not calculate PDN code"}), 400

        logger.info("recalculate_pdn PDN code: %s for user %s, needs_verification: %s", pdn_code, email, needs_verification)

        csv_handler = UserMetadataHandler()
        updated_by = "Admin"

        if not csv_handler.update_pdn_code_with_comment(email, pdn_code, updated_by):
            return jsonify({"error": "Failed to update CSV with new PDN code"}), 500

        logger.info("Successfully updated CSV with PDN code %s for %s by %s", pdn_code, email, updated_by)

        user_data = csv_handler.get_user_by_email(email)
        response_data = {
            "success": True,
            "message": f"PDN code recalculated successfully for {email}",
            "pdn_code": pdn_code,
            "date": user_data.get("Date", "") if user_data else "",
            "updated_by": updated_by,
            "pdn_update_comments": user_data.get("PDN Update Comments", "") if user_data else "",
            "needs_verification": needs_verification,
            "confidence_score": confidence_score
        }

        if calculation_details:
            response_data["calculation_details"] = calculation_details

        return jsonify(response_data)

    except Exception as e:
        logger.error("Error recalculating PDN code, error: %s", e)
        return jsonify({"error": f"Error recalculating PDN code: {str(e)}"}), 500



@pdn_admin_bp.route('/audio/<path:file_path>')
def serve_audio(file_path):
    """Serve audio files with authentication."""

    # Extract session token from query parameters
    session_token = request.args.get('session_token')
    logger.debug("Session token: %s", session_token)

    try:
        verify_session(request.args.get('session_token'))
    except Exception as e:
        logger.error("Session verification failed: %s", e)
        return jsonify({"error": "Unauthorized"}), 401

    # Use the environment variable for saved_results directory
    saved_results_dir = os.getenv('SAVED_RESULTS_DIR', 'saved_results')

    # Handle the file path correctly
    # If the file_path already starts with the saved_results directory structure, use it as is
    if file_path.startswith('pdn/saved_results/'):
        # Remove the 'pdn/saved_results/' prefix and use the environment variable
        relative_path = file_path.replace('pdn/saved_results/', '')
        audio_path = Path(saved_results_dir) / relative_path
    elif file_path.startswith('saved_results/'):
        # Remove the 'saved_results/' prefix
        relative_path = file_path.replace('saved_results/', '')
        audio_path = Path(saved_results_dir) / relative_path
    else:
        # Use the file_path as is
        audio_path = Path(saved_results_dir) / file_path

    logger.debug("Looking for file at: %s", audio_path.absolute())

    # Security check: ensure the path is within the allowed directory
    try:
        audio_path = audio_path.resolve()
        saved_results_path = Path(saved_results_dir).resolve()
        if not audio_path.is_relative_to(saved_results_path):
            logger.warning("Path traversal attempt detected")
            abort(403, description="Access denied")
    except Exception as e:
        logger.error("Path resolution error: %s", e)
        abort(400, description="Invalid file path")

    # Check if file exists (try .mp3 fallback if .wav not found)
    if not audio_path.exists():
        # Try MP3 version if WAV was requested
        if audio_path.suffix.lower() == '.wav':
            mp3_path = audio_path.with_suffix('.mp3')
            if mp3_path.exists():
                audio_path = mp3_path
            else:
                logger.warning("File not found: %s (also tried .mp3)", audio_path)
                abort(404, description="Audio file not found")
        else:
            logger.warning("File not found: %s", audio_path)
            abort(404, description="Audio file not found")

    logger.debug("File found, serving: %s", audio_path)

    try:
        # Detect mimetype based on file extension
        ext = audio_path.suffix.lower()
        mimetypes = {'.wav': 'audio/wav', '.mp3': 'audio/mpeg', '.webm': 'audio/webm', '.ogg': 'audio/ogg'}
        mimetype = mimetypes.get(ext, 'audio/wav')

        return send_file(
            audio_path,
            mimetype=mimetype,
            as_attachment=False,
            download_name=audio_path.name
        )
    except Exception as e:
        logger.error("Error serving audio file: %s", e)
        abort(500, description="Error serving audio file")


@pdn_admin_bp.route('/api/save-audio', methods=['POST'])
def save_audio():
    """Save uploaded audio file."""
    
    try:
        # Check if audio file is present
        if 'audio' not in request.files:
            logger.warning("No audio file in request")
            return jsonify({"error": "No audio file provided"}), 400
        
        audio_file = request.files['audio']
        username = request.form.get('username', 'unknown')
        
        if audio_file.filename == '':
            logger.warning("No audio file selected")
            return jsonify({"error": "No audio file selected"}), 400
        
        # Get the filename from the form data
        filename = audio_file.filename
        
        # Create user directory if it doesn't exist
        pdn_file_path = PDNFilePath()
        user_dir = pdn_file_path.get_user_dir(username)
        user_dir.mkdir(parents=True, exist_ok=True)
        
        # Save the audio file
        audio_path = user_dir / filename
        audio_file.save(audio_path)
        
        logger.info("Audio file saved successfully: %s", audio_path)
        
        return jsonify({
            "message": "Audio saved successfully",
            "filename": filename,
            "path": str(audio_path),
            "status": "success"
        })
        
    except Exception as e:
        logger.error("Error saving audio file: %s", str(e))
        return jsonify({"error": f"Failed to save audio: {str(e)}"}), 500


@pdn_admin_bp.route('/conversation-stats')
def get_conversation_stats():
    """Get conversation statistics for all users"""
    try:
        days = int(request.args.get('days', 7))
        return jsonify({
            "stats": conversation_stats.get_all_stats(days),
            "days": days
        })
    except Exception:
        return jsonify({"error": "Unauthorized"}), 401


@pdn_admin_bp.route('/version')
def get_version():
    """Get application version and release notes"""
    return jsonify({
        "version": VERSION,
        "release_date": RELEASE_DATE,
        "release_notes": RELEASE_NOTES
    })


@pdn_admin_bp.route('/token-usage')
def get_token_usage():
    """Get token usage and cost stats for all Binat chat users"""
    try:
        verify_session(request.args.get('session_token'))
    except Exception as e:
        logger.error("Session verification failed: %s", e)

    try:
        days = int(request.args.get('days', 14))
        from ..pdn_chat_ai.chat_routes import get_agent_instance
        agent = get_agent_instance()
        usage_stats = agent.get_usage_stats(days=days)
        return jsonify({"stats": usage_stats})
    except Exception as e:
        logger.error("Error getting token usage: %s", e)
        return jsonify({"error": str(e)}), 500


@pdn_admin_bp.route('/user/journey/<email>')
def get_user_journey(email):
    """Get user journey timeline data"""
    try:
        verify_session(request.args.get('session_token'))
    except Exception as e:
        logger.error("Session verification failed: %s", e)
        return jsonify({"error": str(e)}), 401

    try:
        # Get user metadata
        csv_handler = UserMetadataHandler()
        user_data = csv_handler.get_user_by_email(email)
        
        # Get conversation stats
        stats = conversation_stats._read_locked()
        
        # Get token usage
        token_data = {}
        try:
            from ..pdn_chat_ai.chat_routes import get_agent_instance
            agent = get_agent_instance()
            token_data = agent.token_usage.get(email, {})
            # Also check by user name
            if not token_data and user_data:
                first_name = user_data.get('First Name', '')
                if first_name:
                    token_data = agent.token_usage.get(first_name, {})
        except Exception as e:
            logger.debug("Could not load token usage for journey: %s", e)
        
        # Build timeline events
        events = []
        
        # Diagnosis event
        if user_data:
            diagnosis_date = user_data.get('Date', '')
            if diagnosis_date and diagnosis_date != 'N/A':
                events.append({
                    'type': 'diagnosis',
                    'date': diagnosis_date,
                    'label': 'אבחון PDN',
                    'detail': f"קוד: {user_data.get('PDN Code', 'N/A')}"
                })
        
        # Conversation events from stats (aggregate per day)
        user_conversations = {}
        for date_str, day_data in stats.items():
            count = day_data.get(email, 0)
            if count > 0:
                user_conversations[date_str] = count
                events.append({
                    'type': 'conversation',
                    'date': date_str,
                    'label': f'{count} שיחות בינת',
                    'detail': f"תאריך: {date_str}"
                })
        
        # Token usage events
        for date_str, day_data in token_data.items():
            if isinstance(day_data, dict) and day_data.get('calls', 0) > 0:
                events.append({
                    'type': 'binat_usage',
                    'date': date_str,
                    'label': f"{day_data['calls']} קריאות AI",
                    'detail': f"טוקנים: {day_data.get('input_tokens', 0) + day_data.get('output_tokens', 0)}"
                })
        
        # Sort events by date
        events.sort(key=lambda e: e['date'], reverse=True)
        
        # Calculate engagement metrics
        total_conversations = sum(user_conversations.values())
        active_days = len(user_conversations)
        
        # Days since diagnosis
        days_since_diagnosis = None
        if user_data and user_data.get('Date'):
            try:
                parts = user_data['Date'].split('/')
                if len(parts) == 3:
                    diag_date = datetime(int(parts[2]), int(parts[1]), int(parts[0]))
                    days_since_diagnosis = (datetime.now() - diag_date).days
            except Exception:
                pass
        
        return jsonify({
            'email': email,
            'user_name': f"{user_data.get('First Name', '')} {user_data.get('Last Name', '')}".strip() if user_data else email,
            'pdn_code': user_data.get('PDN Code', 'N/A') if user_data else 'N/A',
            'events': events,
            'metrics': {
                'total_conversations': total_conversations,
                'active_days': active_days,
                'days_since_diagnosis': days_since_diagnosis,
                'avg_conversations_per_active_day': round(total_conversations / active_days, 1) if active_days > 0 else 0
            }
        })
    except Exception as e:
        logger.error("Error getting user journey: %s", e)
        return jsonify({"error": str(e)}), 500

@pdn_admin_bp.route('/users', methods=['GET'])
def list_users():
    """GET /pdn-admin/users — return all users without passwords."""
    verify_session(request.args.get('session_token'))
    um = get_user_manager()
    return jsonify({"users": um.get_all_users()})


@pdn_admin_bp.route('/users/pdn-codes', methods=['GET'])
def get_pdn_codes():
    """GET /pdn-admin/users/pdn-codes — return available PDN codes from prompt files."""
    verify_session(request.args.get('session_token'))
    um = get_user_manager()
    return jsonify({"codes": um.get_available_pdn_codes()})


@pdn_admin_bp.route('/users', methods=['POST'])
def create_user():
    """POST /pdn-admin/users — add a new user."""
    verify_session(request.args.get('session_token'))

    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    admin_password = data.get('admin_password', '')
    if admin_password.lower() != current_app.config.get('ADMIN_PASSWORD', 'pdn').lower():
        return jsonify({"error": "Invalid admin password"}), 401

    email = data.get('email', '').strip().lower()
    password = data.get('password', '').strip()
    name = data.get('name', '').strip()
    gender = data.get('gender', '').strip()
    pdn_code = data.get('pdn_code', '').strip()
    daily_limit = data.get('daily_conversation_limit', 15)

    try:
        daily_limit = int(daily_limit)
    except (ValueError, TypeError):
        return jsonify({"error": "daily_conversation_limit must be a number"}), 400

    um = get_user_manager()
    try:
        user = um.add_user(email, password, name, pdn_code, daily_limit, gender=gender)
        return jsonify({"success": True, "user": user}), 201
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@pdn_admin_bp.route('/users/<email>', methods=['PUT'])
def update_user_endpoint(email):
    """PUT /pdn-admin/users/<email> — update an existing user."""
    verify_session(request.args.get('session_token'))

    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    admin_password = data.get('admin_password', '')
    data.pop('admin_password', None)
    if admin_password.lower() != current_app.config.get('ADMIN_PASSWORD', 'pdn').lower():
        return jsonify({"error": "Invalid admin password"}), 401

    um = get_user_manager()
    allowed = {'password', 'name', 'gender', 'pdn_code', 'daily_conversation_limit'}
    updates = {k: v for k, v in data.items() if k in allowed and v is not None}

    if 'daily_conversation_limit' in updates:
        try:
            updates['daily_conversation_limit'] = int(updates['daily_conversation_limit'])
        except (ValueError, TypeError):
            return jsonify({"error": "daily_conversation_limit must be a number"}), 400

    try:
        user = um.update_user(email, **updates)
        return jsonify({"success": True, "user": user})
    except KeyError:
        return jsonify({"error": "User not found"}), 404
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@pdn_admin_bp.route('/users/<email>', methods=['DELETE'])
def delete_user_endpoint(email):
    """DELETE /pdn-admin/users/<email> — remove a user."""
    verify_session(request.args.get('session_token'))

    data = request.get_json() or {}
    admin_password = data.get('admin_password', '')
    if admin_password.lower() != current_app.config.get('ADMIN_PASSWORD', 'pdn').lower():
        return jsonify({"error": "Invalid admin password"}), 401

    um = get_user_manager()
    try:
        um.delete_user(email)
        return jsonify({"success": True, "message": f"User {email} deleted"})
    except KeyError:
        return jsonify({"error": "User not found"}), 404


@pdn_admin_bp.route('/download-json')
def download_user_json():
    """Download user's JSON answers file."""

    # Extract session token from query parameters
    session_token = request.args.get('session_token')
    if not session_token:
        logger.warning("No session token provided")
        abort(401, description="No session token provided")

    try:
        verify_session(request.args.get('session_token'))
    except Exception as e:
        logger.error("Session verification failed: %s", e)
        return jsonify({"error": "Unauthorized"}), 401

    # Get file path from query parameters
    file_path = request.args.get('file_path')
    if not file_path:
        logger.warning("No file path provided")
        abort(400, description="No file path provided")

    # Use the environment variable for saved_results directory
    saved_results_dir = os.getenv('SAVED_RESULTS_DIR', 'saved_results')

    # Construct the full file path
    if file_path.startswith('saved_results/'):
        # Remove the 'saved_results/' prefix
        relative_path = file_path.replace('saved_results/', '')
        json_path = Path(saved_results_dir) / relative_path
    else:
        # Use the file_path as is
        json_path = Path(saved_results_dir) / file_path

    logger.debug("Looking for JSON file at: %s", json_path.absolute())

    # Security check: ensure the path is within the allowed directory
    try:
        json_path = json_path.resolve()
        saved_results_path = Path(saved_results_dir).resolve()
        if not json_path.is_relative_to(saved_results_path):
            logger.warning("Path traversal attempt detected")
            abort(403, description="Access denied")
    except Exception as e:
        logger.error("Path resolution error: %s", e)
        abort(400, description="Invalid file path")

    # Check if file exists
    if not json_path.exists():
        logger.warning("JSON file not found: %s", json_path)
        abort(404, description="JSON file not found")

    # Check if it's a JSON file
    if not json_path.suffix.lower() == '.json':
        logger.warning("File is not a JSON file: %s", json_path)
        abort(400, description="File is not a JSON file")

    logger.debug("JSON file found, serving: %s", json_path)

    try:
        return send_file(
            json_path,
            mimetype='application/json',
            as_attachment=True,
            download_name=json_path.name
        )
    except Exception as e:
        logger.error("Error serving JSON file: %s", e)
        abort(500, description="Error serving JSON file")


# --- Coupon Management Routes ---

@pdn_admin_bp.route('/coupons', methods=['GET'])
@require_admin_session
def list_coupons():
    """GET /pdn-admin/coupons — list all coupons with status."""
    try:
        cm = get_coupon_manager()
        coupons = cm.get_all_coupons()
        return jsonify({"coupons": coupons})
    except Exception as e:
        logger.error("Error listing coupons: %s", e)
        return jsonify({"error": "Internal server error"}), 500


@pdn_admin_bp.route('/coupons', methods=['POST'])
@require_admin_session
def create_coupon():
    """POST /pdn-admin/coupons — create a new coupon."""

    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    name = data.get('name', '').strip()
    max_usage = data.get('max_usage')
    code = data.get('code')

    if not name:
        return jsonify({"error": "Name is required"}), 400

    if max_usage is None:
        return jsonify({"error": "max_usage is required"}), 400

    try:
        max_usage = int(max_usage)
    except (ValueError, TypeError):
        return jsonify({"error": "Max usage must be at least 1"}), 400

    if max_usage < 1:
        return jsonify({"error": "Max usage must be at least 1"}), 400

    # Normalize optional code
    if code is not None:
        code = code.strip()
        if not code:
            code = None

    try:
        cm = get_coupon_manager()
        coupon = cm.create_coupon(name, max_usage, code=code)
        return jsonify({"success": True, "coupon": cm.to_response(coupon)}), 201
    except ValueError as e:
        error_msg = str(e)
        if "already exists" in error_msg:
            return jsonify({"error": "Coupon code already exists"}), 409
        if "alphanumeric" in error_msg or "4-20" in error_msg:
            return jsonify({"error": "Code must be 4-20 alphanumeric characters"}), 400
        return jsonify({"error": error_msg}), 400
    except Exception as e:
        logger.error("Error creating coupon: %s", e)
        return jsonify({"error": "Internal server error"}), 500


@pdn_admin_bp.route('/coupons/<code>', methods=['PUT'])
@require_admin_session
def update_coupon(code):
    """PUT /pdn-admin/coupons/<code> — update a coupon (name, max_usage)."""
    if not _validate_coupon_code_param(code):
        return jsonify({"error": "Invalid coupon code format"}), 400

    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    updates = {}
    if 'name' in data:
        updates['name'] = data['name'].strip() if isinstance(data['name'], str) else data['name']
    if 'max_usage' in data:
        try:
            updates['max_usage'] = int(data['max_usage'])
        except (ValueError, TypeError):
            return jsonify({"error": "Max usage must be at least 1"}), 400
        if updates['max_usage'] < 1:
            return jsonify({"error": "Max usage must be at least 1"}), 400

    if not updates:
        return jsonify({"error": "No valid fields to update"}), 400

    try:
        cm = get_coupon_manager()
        coupon = cm.update_coupon(code, **updates)
        return jsonify({"success": True, "coupon": cm.to_response(coupon)})
    except KeyError:
        return jsonify({"error": "Coupon not found"}), 404
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logger.error("Error updating coupon: %s", e)
        return jsonify({"error": "Internal server error"}), 500


@pdn_admin_bp.route('/coupons/<code>', methods=['DELETE'])
@require_admin_session
def delete_coupon(code):
    """DELETE /pdn-admin/coupons/<code> — delete a coupon."""
    if not _validate_coupon_code_param(code):
        return jsonify({"error": "Invalid coupon code format"}), 400

    try:
        cm = get_coupon_manager()
        cm.delete_coupon(code)
        return jsonify({"success": True, "message": f"Coupon {code} deleted"})
    except KeyError:
        return jsonify({"error": "Coupon not found"}), 404
    except Exception as e:
        logger.error("Error deleting coupon: %s", e)
        return jsonify({"error": "Internal server error"}), 500


@pdn_admin_bp.route('/coupons/<code>/usage', methods=['GET'])
@require_admin_session
def get_coupon_usage(code):
    """GET /pdn-admin/coupons/<code>/usage — get usage details (used_by list)."""
    if not _validate_coupon_code_param(code):
        return jsonify({"error": "Invalid coupon code format"}), 400

    try:
        cm = get_coupon_manager()
        coupon = cm.get_coupon(code)
        if coupon is None:
            return jsonify({"error": "Coupon not found"}), 404
        response = cm.to_response(coupon)
        return jsonify(response)
    except Exception as e:
        logger.error("Error getting coupon usage: %s", e)
        return jsonify({"error": "Internal server error"}), 500


@pdn_admin_bp.route('/coupons/<code>/send-invite', methods=['POST'])
@require_admin_session
def send_coupon_invite(code):
    """POST /pdn-admin/coupons/<code>/send-invite — send coupon invite email."""
    if not _validate_coupon_code_param(code):
        return jsonify({"error": "Invalid coupon code format"}), 400

    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    recipient_email = data.get('email', '').strip().lower()
    if not recipient_email or '@' not in recipient_email:
        return jsonify({"error": "Valid email address is required"}), 400

    try:
        cm = get_coupon_manager()
        coupon = cm.get_coupon(code)
        if coupon is None:
            return jsonify({"error": "Coupon not found"}), 404

        # Determine base URL from request
        base_url = request.host_url.rstrip('/')

        if not send_coupon_invite_email(recipient_email, code, base_url):
            return jsonify({"error": "Failed to send email"}), 500

        return jsonify({
            "success": True,
            "message": f"Coupon invite sent to {recipient_email}"
        })
    except Exception as e:
        logger.error("Error sending coupon invite: %s", e)
        return jsonify({"error": "Internal server error"}), 500


# --- PDN Code Analysis Report ---

_PDN_12_CODES_LIST = ["E1", "E5", "E9", "A3", "A7", "A11", "T4", "T8", "T12", "P2", "P6", "P10"]
_PDN_12_CODES_SET = frozenset(_PDN_12_CODES_LIST)
_APP_DATA_DIR = Path(__file__).resolve().parent.parent / 'data'

# Cache for questions and test emails (read once per process, thread-safe via GIL for reads).
# NOTE: Cache invalidates only on process restart. If questions.json or test_users.json change,
# restart the Gunicorn workers to pick up new data.
_cache_lock = threading.Lock()
_questions_cache = None
_test_emails_cache = None


def _load_test_emails():
    """Load test user emails from config file. Thread-safe cached after first load."""
    global _test_emails_cache
    if _test_emails_cache is not None:
        return _test_emails_cache

    with _cache_lock:
        # Double-check after acquiring lock
        if _test_emails_cache is not None:
            return _test_emails_cache

        test_users_file = _APP_DATA_DIR / 'test_users.json'
        if not test_users_file.exists():
            _test_emails_cache = frozenset()
            return _test_emails_cache
        try:
            with open(test_users_file, 'r', encoding='utf-8') as f:
                test_data = json.load(f)
            _test_emails_cache = frozenset(e.lower() for e in test_data.get('test_emails', []))
        except (json.JSONDecodeError, IOError) as e:
            logger.warning("Failed to load test_users.json: %s", e)
            _test_emails_cache = frozenset()
    return _test_emails_cache


def _load_questions_cached():
    """Load and cache questions (excluding PersonalDetails and PartF). Thread-safe."""
    global _questions_cache
    if _questions_cache is not None:
        return _questions_cache

    with _cache_lock:
        # Double-check after acquiring lock
        if _questions_cache is not None:
            return _questions_cache

        questions_file = _APP_DATA_DIR / 'questions.json'
        questions = {}
        try:
            with open(questions_file, 'r', encoding='utf-8') as f:
                q_data = json.load(f)
            for phase_name, phase_data in q_data.get('phases', {}).items():
                if phase_name in ('PersonalDetails', 'PartF'):
                    continue
                for q_num, q_info in phase_data.get('questions', {}).items():
                    codes = set()
                    for opt in q_info.get('options', []):
                        if opt.get('code'):
                            codes.add(opt['code'])
                    questions[q_num] = {
                        'text': q_info.get('text', ''),
                        'codes': '/'.join(sorted(codes)) if codes else '-',
                        'phase': phase_name,
                        'options': q_info.get('options', []),
                    }
        except (json.JSONDecodeError, IOError) as e:
            logger.error("Failed to load questions.json: %s", e)
        _questions_cache = questions
    return _questions_cache


@pdn_admin_bp.route('/pdn-analysis')
def pdn_analysis_page():
    """PDN Code Analysis report page. Restricted to authorized admin only."""
    return render_template("pdn_analysis.html")


_PDN_ANALYSIS_ALLOWED_EMAILS = frozenset(['tomergur@gmail.com'])


@pdn_admin_bp.route('/api/pdn-analysis/data', methods=['GET', 'POST'])
def pdn_analysis_data():
    """
    Get all data needed for PDN analysis report.

    Returns JSON with: users (filtered), answers (per user), questions, pdn_codes.
    Accepts optional 'emails' filter (POST body or query param) to limit data loading.
    Restricted to authorized admins only.

    NOTE: All user-facing text in answers comes from controlled JSON files (questions.json)
    authored by admins. The trust boundary is the admin panel itself.
    """
    # Support token from Authorization header, POST body, or query param (fallback)
    token = None
    auth_header = request.headers.get('Authorization', '')
    if auth_header.startswith('Bearer '):
        token = auth_header[7:]
    elif request.is_json and request.json:
        token = request.json.get('session_token')
    if not token:
        token = request.args.get('session_token')

    try:
        session = verify_session(token)
    except HTTPException:
        return jsonify({"error": "Unauthorized"}), 401

    # Check if this admin is allowed to access PDN analysis
    session_email = (session.get('email') or '').lower()
    if session_email not in _PDN_ANALYSIS_ALLOWED_EMAILS:
        return jsonify({"error": "Access denied"}), 403

    try:
        # Get email filter (from POST body or query param)
        emails_filter = None
        if request.is_json and request.json:
            emails_raw = request.json.get('emails')
        else:
            emails_raw = request.args.get('emails')
        if emails_raw:
            emails_filter = frozenset(e.strip().lower() for e in emails_raw.split(',') if e.strip())

        csv_metadata_handler = UserMetadataHandler()
        users_metadata = load_user_metadata()
        test_emails = _load_test_emails()

        # Filter: only valid diagnosed PDN codes (human-validated), exclude test users
        valid_users = []
        for u in users_metadata:
            email = u.get('email', '').strip()
            # Use diagnose_pdn_code (human-validated) instead of pdn_code (system calc)
            pdn_code = u.get('diagnose_pdn_code', '').strip()
            if pdn_code not in _PDN_12_CODES_SET:
                continue
            if email.lower() in test_emails:
                continue
            if emails_filter and email.lower() not in emails_filter:
                continue
            valid_users.append({
                'uid': u.get('user_id', ''),
                'email': email,
                'pdn_code': pdn_code,
                'first_name': u.get('first_name', ''),
                'last_name': u.get('last_name', ''),
            })

        # Fetch answers - only selected_option_code and ranking (no question_options)
        user_answers = {}
        for user in valid_users:
            email = user['email']
            try:
                answers = csv_metadata_handler.get_user_files(email, "answers")
                if answers and isinstance(answers, dict):
                    filtered = {}
                    for k, v in answers.items():
                        if k == 'metadata':
                            continue
                        if not isinstance(v, dict):
                            continue
                        if 'selected_option_code' in v or 'ranking' in v:
                            entry = {}
                            if 'selected_option_code' in v:
                                entry['selected_option_code'] = v['selected_option_code']
                            if 'ranking' in v:
                                entry['ranking'] = v['ranking']
                            filtered[k] = entry
                    if filtered:
                        user_answers[email] = filtered
            except Exception as e:
                logger.debug("Could not load answers for %s: %s", email, e)

        questions = _load_questions_cached()

        response = jsonify({
            'users': valid_users,
            'answers': user_answers,
            'questions': questions,
            'pdn_codes': _PDN_12_CODES_LIST,
        })
        response.headers['Cache-Control'] = 'private, max-age=60'
        return response

    except Exception as e:
        logger.error("Error in PDN analysis: %s", e, exc_info=True)
        return jsonify({"error": "Internal server error"}), 500


@pdn_admin_bp.route('/api/pdn-analysis/excel', methods=['GET', 'POST'])
def pdn_analysis_excel():
    """
    Generate and stream an Excel (.xlsx) file for the PDN analysis report.
    Same auth/filter logic as pdn_analysis_data.
    Restricted to authorized admins only.
    """
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl is not installed on this server"}), 500

    # --- Auth ---
    token = None
    auth_header = request.headers.get('Authorization', '')
    if auth_header.startswith('Bearer '):
        token = auth_header[7:]
    elif request.is_json and request.json:
        token = request.json.get('session_token')
    if not token:
        token = request.args.get('session_token')

    try:
        session = verify_session(token)
    except Exception:
        return jsonify({"error": "Unauthorized"}), 401

    session_email = (session.get('email') or '').lower()
    if session_email not in _PDN_ANALYSIS_ALLOWED_EMAILS:
        return jsonify({"error": "Access denied"}), 403

    try:
        # --- Load same data as pdn_analysis_data ---
        emails_filter = None
        if request.is_json and request.json:
            emails_raw = request.json.get('emails')
        else:
            emails_raw = request.args.get('emails')
        if emails_raw:
            emails_filter = frozenset(e.strip().lower() for e in emails_raw.split(',') if e.strip())

        csv_metadata_handler = UserMetadataHandler()
        users_metadata = load_user_metadata()
        test_emails = _load_test_emails()

        valid_users = []
        for u in users_metadata:
            email = u.get('email', '').strip()
            pdn_code = u.get('diagnose_pdn_code', '').strip()
            if pdn_code not in _PDN_12_CODES_SET:
                continue
            if email.lower() in test_emails:
                continue
            if emails_filter and email.lower() not in emails_filter:
                continue
            valid_users.append({
                'uid': u.get('user_id', ''),
                'email': email,
                'pdn_code': pdn_code,
                'first_name': u.get('first_name', ''),
                'last_name': u.get('last_name', ''),
            })

        user_answers = {}
        for user in valid_users:
            email = user['email']
            try:
                answers = csv_metadata_handler.get_user_files(email, "answers")
                if answers and isinstance(answers, dict):
                    filtered = {}
                    for k, v in answers.items():
                        if k == 'metadata':
                            continue
                        if not isinstance(v, dict):
                            continue
                        if 'selected_option_code' in v or 'ranking' in v:
                            entry = {}
                            if 'selected_option_code' in v:
                                entry['selected_option_code'] = v['selected_option_code']
                            if 'ranking' in v:
                                entry['ranking'] = v['ranking']
                            filtered[k] = entry
                    if filtered:
                        user_answers[email] = filtered
            except Exception:
                pass

        questions_data = _load_questions_cached()

        # --- Build workbook ---
        PDN_12 = ["E1", "E5", "E9", "A3", "A7", "A11", "T4", "T8", "T12", "P2", "P6", "P10"]

        def _fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def _font(bold=False, color="000000", size=10):
            return Font(bold=bold, color=color, size=size)

        def _border():
            thin = Side(style="thin", color="CCCCCC")
            return Border(left=thin, right=thin, top=thin, bottom=thin)

        CODE_FILLS = {"E": _fill("D6EAF8"), "A": _fill("D4EFDF"), "T": _fill("FDEBD0"), "P": _fill("FADBD8")}
        ANSWER_FILLS = {
            "AP": _fill("D4EFDF"), "ET": _fill("D6EAF8"), "AE": _fill("FADBD8"), "TP": _fill("FDEBD0"),
            "A": _fill("D4EFDF"), "E": _fill("D6EAF8"), "P": _fill("FADBD8"), "T": _fill("FDEBD0"),
        }
        HEADER_FILL = _fill("2C3E50")
        HEADER_FONT = _font(bold=True, color="FFFFFF", size=10)
        PHASE_FILL  = _fill("7F8C8D")
        PHASE_FONT  = _font(bold=True, color="FFFFFF", size=10)
        CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
        RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=True)

        def _get_dominant_answer(answer_entry):
            """Return the dominant answer code from an answer entry."""
            if not isinstance(answer_entry, dict):
                return None
            if 'selected_option_code' in answer_entry:
                return answer_entry['selected_option_code']
            if 'ranking' in answer_entry:
                ranking = answer_entry['ranking']
                if isinstance(ranking, dict) and ranking:
                    return max(ranking, key=ranking.get)
            return None

        # Sort questions numerically
        def _sort_key(q):
            try:
                return int(q)
            except ValueError:
                return 999

        sorted_phases = ["PartA", "PartB", "PartC", "PartD", "PartE"]
        phase_labels = {
            "PartA": "חלק א - בחירה בינארית",
            "PartB": "חלק ב - דירוג",
            "PartC": "חלק ג - סולם",
            "PartD": "חלק ד - ילדות/בגרות",
            "PartE": "חלק ה - דירוג 4",
        }

        # Build sorted question list from questions_data dict
        # questions_data: {q_num: {text, codes, phase, options}}
        sorted_q_nums = sorted(questions_data.keys(), key=_sort_key)

        # Group users by PDN code
        users_by_code = {}
        for code in PDN_12:
            users_by_code[code] = [u for u in valid_users if u['pdn_code'] == code]

        # Compute stats: for each question, for each code, count dominant answers
        from collections import defaultdict
        stats = {}
        for q_num in sorted_q_nums:
            stats[q_num] = {}
            for code in PDN_12:
                counts = defaultdict(int)
                total = 0
                for user in users_by_code[code]:
                    ans = user_answers.get(user['email'], {}).get(str(q_num))
                    if ans:
                        dom = _get_dominant_answer(ans)
                        if dom:
                            counts[dom] += 1
                            total += 1
                stats[q_num][code] = {"counts": dict(counts), "total": total}

        wb = openpyxl.Workbook()

        # ---- Sheet 1: User Answers ----
        ws1 = wb.active
        ws1.title = "תשובות משתמשים"
        ws1.sheet_view.rightToLeft = True

        ws1.cell(1, 1, "משתמש (PDN | UID | שם)").fill = HEADER_FILL
        ws1.cell(1, 1).font = HEADER_FONT
        ws1.cell(1, 1).alignment = RIGHT
        ws1.column_dimensions["A"].width = 28

        for col_idx, q_num in enumerate(sorted_q_nums, start=2):
            c = ws1.cell(1, col_idx, f"Q{q_num}")
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = CENTER
            c.border = _border()
            ws1.column_dimensions[get_column_letter(col_idx)].width = 7

        ws1.row_dimensions[1].height = 30
        ws1.freeze_panes = "B2"

        for row_idx, user in enumerate(valid_users, start=2):
            email  = user['email']
            answers = user_answers.get(email, {})
            pdn    = user['pdn_code'] or "NA"
            name   = f"{user.get('first_name','')} {user.get('last_name','')}".strip()
            label  = f"{pdn} | {user['uid']}" + (f" | {name}" if name else "")

            uc = ws1.cell(row_idx, 1, label)
            uc.alignment = RIGHT
            uc.border = _border()
            if pdn and pdn[0] in CODE_FILLS:
                uc.fill = CODE_FILLS[pdn[0]]

            for col_idx, q_num in enumerate(sorted_q_nums, start=2):
                ans = answers.get(str(q_num))
                val = _get_dominant_answer(ans) if ans else None
                c = ws1.cell(row_idx, col_idx, val or "")
                c.alignment = CENTER
                c.border = _border()
                if val and val in ANSWER_FILLS:
                    c.fill = ANSWER_FILLS[val]

        # ---- Sheet 2: Stats per PDN code ----
        ws2 = wb.create_sheet("סטטיסטיקה לפי קוד")
        ws2.sheet_view.rightToLeft = True

        ws2.cell(1, 1, "שאלה").fill = HEADER_FILL
        ws2.cell(1, 1).font = HEADER_FONT
        ws2.cell(1, 1).alignment = RIGHT
        ws2.column_dimensions["A"].width = 40

        for col_idx, code in enumerate(PDN_12, start=2):
            count = len(users_by_code.get(code, []))
            c = ws2.cell(1, col_idx, f"{code}\n({count})")
            c.fill = CODE_FILLS.get(code[0], HEADER_FILL)
            c.font = _font(bold=True, color="000000", size=10)
            c.alignment = CENTER
            c.border = _border()
            ws2.column_dimensions[get_column_letter(col_idx)].width = 16

        ws2.row_dimensions[1].height = 36
        ws2.freeze_panes = "B2"

        row_idx = 2
        current_phase = None
        for q_num in sorted_q_nums:
            q = questions_data.get(q_num, {})
            phase = q.get('phase', '')
            if phase != current_phase:
                current_phase = phase
                label = phase_labels.get(phase, phase)
                for col in range(1, len(PDN_12) + 2):
                    c = ws2.cell(row_idx, col, label if col == 1 else "")
                    c.fill = PHASE_FILL
                    c.font = PHASE_FONT
                    c.alignment = CENTER
                    c.border = _border()
                ws2.row_dimensions[row_idx].height = 18
                row_idx += 1

            codes_str = q.get('codes', '')
            short_text = (q.get('text', '') or '')[:55]
            lc = ws2.cell(row_idx, 1, f"Q{q_num} [{codes_str}]\n{short_text}")
            lc.alignment = RIGHT
            lc.border = _border()
            ws2.row_dimensions[row_idx].height = 36

            for col_idx, code in enumerate(PDN_12, start=2):
                s = stats[q_num].get(code, {"counts": {}, "total": 0})
                total = s["total"]
                if total == 0:
                    c = ws2.cell(row_idx, col_idx, "-")
                    c.alignment = CENTER
                    c.border = _border()
                    col_idx += 1
                    continue
                sorted_ans = sorted(s["counts"].items(), key=lambda x: -x[1])
                lines = [f"{a}: {round(n/total*100)}%" for a, n in sorted_ans[:3]]
                c = ws2.cell(row_idx, col_idx, "\n".join(lines))
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = _border()
                if sorted_ans and sorted_ans[0][0] in ANSWER_FILLS:
                    c.fill = ANSWER_FILLS[sorted_ans[0][0]]

            row_idx += 1

        # ---- Sheet 3: Summary ----
        ws3 = wb.create_sheet("סיכום קודים")
        ws3.sheet_view.rightToLeft = True
        ws3.cell(1, 1, "קוד PDN").fill = HEADER_FILL
        ws3.cell(1, 1).font = HEADER_FONT
        ws3.cell(1, 1).alignment = CENTER
        ws3.cell(1, 2, "מספר מאובחנים").fill = HEADER_FILL
        ws3.cell(1, 2).font = HEADER_FONT
        ws3.cell(1, 2).alignment = CENTER
        ws3.column_dimensions["A"].width = 14
        ws3.column_dimensions["B"].width = 20

        for r, code in enumerate(PDN_12, start=2):
            count = len(users_by_code.get(code, []))
            c1, c2 = ws3.cell(r, 1, code), ws3.cell(r, 2, count)
            c1.alignment = CENTER
            c2.alignment = CENTER
            c1.border = _border()
            c2.border = _border()
            if code[0] in CODE_FILLS:
                c1.fill = CODE_FILLS[code[0]]
                c2.fill = CODE_FILLS[code[0]]

        ws3.cell(len(PDN_12) + 2, 1, 'סה"כ').font = _font(bold=True)
        ws3.cell(len(PDN_12) + 2, 2, len(valid_users)).font = _font(bold=True)

        # --- Stream response ---
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"pdn_matrix_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response = make_response(buf.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        logger.error("Error generating PDN Excel: %s", e, exc_info=True)
        return jsonify({"error": "Internal server error"}), 500
